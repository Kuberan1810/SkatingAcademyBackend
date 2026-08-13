import calendar
from openpyxl import styles
from datetime import date, datetime
from io import BytesIO, StringIO
import csv

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
)

from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session

from app.auth.dependencies import get_current_admin
from app.core.dependencies import get_db

from app.models.admin import Admin
from app.models.fee import FeePayment
from app.models.student import Student
from app.models.batch import Batch


# =========================================================
# PDF IMPORTS
# =========================================================

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)


# =========================================================
# ROUTER
# =========================================================

router = APIRouter(
    prefix="/fees",
    tags=["Fees"],
)


# =========================================================
# HELPER
# =========================================================

def format_payment_date(
    payment_date,
) -> str:

    if payment_date is None:
        return ""

    if isinstance(
        payment_date,
        datetime,
    ):
        return payment_date.strftime(
            "%d %b %Y %I:%M %p"
        )

    if isinstance(
        payment_date,
        date,
    ):
        return payment_date.strftime(
            "%d %b %Y"
        )

    return str(payment_date)


# =========================================================
# EXPORT FEE REPORT
# =========================================================

@router.get("/export")
def export_fee_report(

    from_date: date | None = Query(
        default=None,
        description="Report start date",
    ),

    to_date: date | None = Query(
        default=None,
        description="Report end date",
    ),

    batch_id: int | None = Query(
        default=None,
        gt=0,
        description="Filter by batch",
    ),

    payment_method: str | None = Query(
        default=None,
        description="CASH / UPI / CARD",
    ),

    format: str = Query(
        default="xlsx",
        pattern="^(xlsx|csv|pdf)$",
        description="xlsx / csv / pdf",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):

    # =====================================================
    # VALIDATE DATES
    # =====================================================

    if (
        from_date is not None
        and to_date is not None
        and from_date > to_date
    ):

        raise HTTPException(
            status_code=400,
            detail=(
                "from_date cannot be greater "
                "than to_date"
            ),
        )

    # =====================================================
    # BASE QUERY
    # =====================================================

    query = (
        db.query(
            FeePayment,
            Student,
            Batch,
        )
        .join(
            Student,
            FeePayment.student_id
            == Student.id,
        )
        .join(
            Batch,
            Student.batch_id
            == Batch.id,
        )
        .filter(
            Student.is_active.is_(True),
            Batch.is_active.is_(True),
        )
    )

    # =====================================================
    # DATE FILTER
    #
    # IMPORTANT:
    # Uses actual payment_date.
    # =====================================================

    if from_date is not None:

        query = query.filter(
            FeePayment.payment_date
            >= datetime.combine(
                from_date,
                datetime.min.time(),
            )
        )

    if to_date is not None:

        # Include the complete to_date
        next_day = (
            datetime.combine(
                to_date,
                datetime.min.time(),
            )
            .replace(
                day=to_date.day,
            )
        )

        # Better exclusive upper boundary
        if to_date == date(
            to_date.year,
            12,
            31,
        ):

            next_day = datetime(
                to_date.year + 1,
                1,
                1,
            )

        else:

            next_day = datetime.combine(
                to_date,
                datetime.min.time(),
            )

            from datetime import timedelta

            next_day = (
                next_day
                + timedelta(days=1)
            )

        query = query.filter(
            FeePayment.payment_date
            < next_day
        )

    # =====================================================
    # BATCH FILTER
    # =====================================================

    if batch_id is not None:

        query = query.filter(
            Student.batch_id
            == batch_id
        )

    # =====================================================
    # PAYMENT METHOD FILTER
    # =====================================================

    if payment_method:

        query = query.filter(
            FeePayment.payment_method
            == payment_method.upper()
        )

    # =====================================================
    # GET PAYMENTS
    # =====================================================

    payments = (
        query
        .order_by(
            FeePayment.payment_date.desc()
        )
        .all()
    )

    # =====================================================
    # REPORT ROWS
    # =====================================================

    rows = []

    total_collected = 0

    for payment, student, batch in payments:

        amount = int(
            payment.net_payable or 0
        )

        total_collected += amount

        # =====================================================
        # FEE MONTH NAME
        # =====================================================

        fee_month = (
            calendar.month_name[
                payment.fee_month
            ]
            if payment.fee_month
            and 1 <= payment.fee_month <= 12
            else ""
        )

        # =====================================================
        # REPORT ROW
        # =====================================================

        rows.append(
            {
                "Payment ID":
                    str(payment.id),

                "Student ID":
                    str(student.id),

                "Student Name":
                    student.full_name,

                "Batch Name":
                    batch.batch_name,

                "Location":
                    batch.location or "",

                "Phone":
                    student.phone_number,

                "Fee Month":
                    fee_month,

                "Fee Year":
                    payment.fee_year,

                "Amount":
                    amount,

                "Payment Method":
                    str(
                        payment.payment_method
                    ),

                "Payment Date":
                    format_payment_date(
                        payment.payment_date
                    ),

                "Status":
                    "PAID",
            }
        )

    # =====================================================
    # NO DATA
    # =====================================================

    if not rows:

        raise HTTPException(
            status_code=404,
            detail=(
                "No fee payments found "
                "for the selected filters"
            ),
        )

    # =====================================================
    # EXCEL EXPORT
    # =====================================================

    if format == "xlsx":

        from openpyxl import Workbook

        from openpyxl.styles import (
            Font,
            Alignment,
        )

        from openpyxl.utils import (
            get_column_letter,
        )

        # -------------------------------------------------
        # CREATE WORKBOOK
        # -------------------------------------------------

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Fee Report"

        # -------------------------------------------------
        # TITLE
        # -------------------------------------------------

        worksheet["A1"] = (
            "FEE COLLECTION REPORT"
        )

        worksheet["A1"].font = Font(
            bold=True,
            size=16,
        )

        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=12,
        )

        worksheet["A1"].alignment = (
            Alignment(
                horizontal="center"
            )
        )

        # -------------------------------------------------
        # FILTER INFORMATION
        # -------------------------------------------------

        worksheet["A2"] = "From Date"

        worksheet["B2"] = (
            from_date.strftime(
                "%d %b %Y"
            )
            if from_date
            else "All"
        )

        worksheet["C2"] = "To Date"

        worksheet["D2"] = (
            to_date.strftime(
                "%d %b %Y"
            )
            if to_date
            else "All"
        )

        worksheet["E2"] = "Batch"

        worksheet["F2"] = (
            str(batch_id)
            if batch_id
            else "All"
        )

        worksheet["G2"] = (
            "Payment Method"
        )

        worksheet["H2"] = (
            payment_method.upper()
            if payment_method
            else "All"
        )

        # -------------------------------------------------
        # SUMMARY
        # -------------------------------------------------

        worksheet["A4"] = (
            "Total Payments"
        )

        worksheet["B4"] = len(rows)

        worksheet["C4"] = (
            "Total Collected"
        )

        worksheet["D4"] = (
            total_collected
        )

        worksheet["C4"].font = Font(
            bold=True
        )

        worksheet["D4"].font = Font(
            bold=True
        )

        # -------------------------------------------------
        # HEADERS
        # -------------------------------------------------

        header_row = 6

        headers = list(
            rows[0].keys()
        )

        for (
            column_index,
            header,
        ) in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )

            cell.font = Font(
                bold=True
            )

            cell.alignment = (
                Alignment(
                    horizontal="center"
                )
            )

        # -------------------------------------------------
        # DATA
        # -------------------------------------------------

        for (
            row_index,
            row_data,
        ) in enumerate(
            rows,
            start=header_row + 1,
        ):

            for (
                column_index,
                header,
            ) in enumerate(
                headers,
                start=1,
            ):

                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=row_data[
                        header
                    ],
                )

                if header == "Amount":

                    cell.number_format = (
                        '₹#,##0'
                    )

        # -------------------------------------------------
        # COLUMN WIDTH
        # -------------------------------------------------

        for (
            column_index,
            header,
        ) in enumerate(
            headers,
            start=1,
        ):

            max_length = len(header)

            for row_index in range(
                header_row + 1,
                worksheet.max_row + 1,
            ):

                value = worksheet.cell(
                    row=row_index,
                    column=column_index,
                ).value

                if value is not None:

                    max_length = max(
                        max_length,
                        len(str(value)),
                    )

            worksheet.column_dimensions[
                get_column_letter(
                    column_index
                )
            ].width = min(
                max_length + 3,
                35,
            )

        # -------------------------------------------------
        # FREEZE HEADER
        # -------------------------------------------------

        worksheet.freeze_panes = "A7"

        # -------------------------------------------------
        # FILTER
        # -------------------------------------------------

        worksheet.auto_filter.ref = (
            f"A6:L{worksheet.max_row}"
        )

        # -------------------------------------------------
        # SAVE
        # -------------------------------------------------

        output = BytesIO()

        workbook.save(output)

        output.seek(0)

        filename = (
            "fee-report-"
            f"{date.today().isoformat()}"
            ".xlsx"
        )

        return StreamingResponse(
            output,
            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition":
                    (
                        "attachment; "
                        f"filename={filename}"
                    )
            },
        )

    # =====================================================
    # PDF EXPORT
    # =====================================================

    if format == "pdf":

        output = BytesIO()

        # -------------------------------------------------
        # PDF DOCUMENT
        # -------------------------------------------------

        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            rightMargin=10 * mm,
            leftMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=10 * mm,
        )

        styles = (
            getSampleStyleSheet()
        )

        title_style = (
            styles["Title"]
        )

        title_style.fontSize = 18

        title_style.leading = 22

        story = []

        # -------------------------------------------------
        # TITLE
        # -------------------------------------------------

        story.append(
            Paragraph(
                "Fee Collection Report",
                title_style,
            )
        )

        story.append(
            Spacer(
                1,
                5 * mm,
            )
        )

        # =========================================================
        # FILTER VALUES
        # =========================================================

        from_text = (
            from_date.strftime(
                "%d %b %Y"
            )
            if from_date
            else "All"
        )

        to_text = (
            to_date.strftime(
                "%d %b %Y"
            )
            if to_date
            else "All"
        )

        # ---------------------------------------------------------
        # BATCH NAME
        # ---------------------------------------------------------

        batch_text = "All"

        if batch_id:

            if payments:

                selected_batch = payments[0][2]

                batch_text = (
                    selected_batch.batch_name
                    or f"Batch {batch_id}"
                )

        # ---------------------------------------------------------
        # PAYMENT METHOD
        # ---------------------------------------------------------

        method_text = (
            payment_method.upper()
            if payment_method
            else "All"
        )
        # -------------------------------------------------
        # SUMMARY
        # -------------------------------------------------

        summary_data = [
            [
                "From Date",
                from_text,
                "To Date",
                to_text,
            ],
            [
                "Batch",
                batch_text,
                "Payment Method",
                method_text,
            ],
            [
                "Total Payments",
                str(len(rows)),
                "Total Collected",
                f"Rs. {total_collected:,.2f}",
            ],
        ]

        summary_table = Table(
            summary_data,
            colWidths=[
                30 * mm,
                45 * mm,
                35 * mm,
                55 * mm,
            ],
        )

        summary_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (0, -1),
                        colors.HexColor(
                            "#F3F4F6"
                        ),
                    ),
                    (
                        "BACKGROUND",
                        (2, 0),
                        (2, -1),
                        colors.HexColor(
                            "#F3F4F6"
                        ),
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, -1),
                        "Helvetica",
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (0, -1),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTNAME",
                        (2, 0),
                        (2, -1),
                        "Helvetica-Bold",
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.grey,
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        6,
                    ),
                ]
            )
        )

        story.append(
            summary_table
        )

        story.append(
            Spacer(
                1,
                7 * mm,
            )
        )

        # -------------------------------------------------
        # PDF TABLE HEADERS
        # -------------------------------------------------

        headers = [
            "Payment ID",
            "Student",
            "Batch",
            "Phone",
            "Fee Month",
            "Fee Year",
            "Amount",
            "Method",
            "Payment Date",
            "Status",
        ]

        table_data = [
            headers
        ]

        # -------------------------------------------------
        # PDF TABLE DATA
        # -------------------------------------------------

        for row in rows:

            table_data.append(
                [
                    row["Payment ID"],
                    row["Student Name"],
                    row["Batch Name"],
                    row["Phone"] or "",
                    row["Fee Month"],
                    row["Fee Year"],
                    f"Rs. {row['Amount']:,.2f}",
                    row["Payment Method"],
                    row["Payment Date"],
                    row["Status"],
                ]

            )

        # -------------------------------------------------
        # REPORT TABLE
        # -------------------------------------------------

        report_table = Table(
            table_data,
            repeatRows=1,
            colWidths=[
                20 * mm,
                38 * mm,
                32 * mm,
                30 * mm,
                18 * mm,
                18 * mm,
                25 * mm,
                25 * mm,
                35 * mm,
                18 * mm,
            ],
        )

        report_table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor(
                            "#111827"
                        ),
                    ),
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white,
                    ),
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),
                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, -1),
                        7,
                    ),
                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.3,
                        colors.grey,
                    ),
                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE",
                    ),
                    (
                        "ALIGN",
                        (4, 1),
                        (6, -1),
                        "CENTER",
                    ),
                    (
                        "ALIGN",
                        (9, 1),
                        (9, -1),
                        "CENTER",
                    ),
                    (
                        "PADDING",
                        (0, 0),
                        (-1, -1),
                        4,
                    ),
                ]
            )
        )

        story.append(
            report_table
        )

        story.append(
            Spacer(
                1,
                6 * mm,
            )
        )

        # -------------------------------------------------
        # FOOTER
        # -------------------------------------------------

        story.append(
            Paragraph(
                (
                    "<b>Total Collected:</b> "
                    f"Rs. {total_collected:,.2f}"
                ),
                styles["Normal"],
            )
        )

        # -------------------------------------------------
        # BUILD PDF
        # -------------------------------------------------

        document.build(
            story
        )

        output.seek(0)

        filename = (
            "fee-report-"
            f"{date.today().isoformat()}"
            ".pdf"
        )

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                    (
                        "attachment; "
                        f"filename={filename}"
                    )
            },
        )

    # =====================================================
    # CSV EXPORT
    # =====================================================

    output = StringIO()

    writer = csv.DictWriter(
        output,
        fieldnames=list(
            rows[0].keys()
        ),
    )

    writer.writeheader()

    writer.writerows(
        rows
    )

    output.seek(0)

    filename = (
        "fee-report-"
        f"{date.today().isoformat()}"
        ".csv"
    )

    return StreamingResponse(
        iter(
            [
                output.getvalue()
            ]
        ),
        media_type="text/csv",
        headers={
            "Content-Disposition":
                (
                    "attachment; "
                    f"filename={filename}"
                )
        },
    )