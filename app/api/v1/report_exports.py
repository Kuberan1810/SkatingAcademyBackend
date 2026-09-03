from datetime import date, datetime, time
import calendar
import csv
from io import BytesIO, StringIO

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
)

from fastapi.responses import StreamingResponse

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.auth.dependencies import get_current_admin
from app.core.dependencies import get_db

from app.models.admin import Admin
from app.models.student import Student
from app.models.batch import Batch
from app.models.session import Session as SessionModel
from app.models.attendance import Attendance
from app.models.fee import FeePayment


router = APIRouter(
    prefix="/reports",
    tags=["Report Exports"],
)


# =========================================================
# HELPERS
# =========================================================


def format_date(value):

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime(
            "%d %b %Y %I:%M %p"
        )

    if isinstance(value, date):
        return value.strftime(
            "%d %b %Y"
        )

    return str(value)


def calculate_age(dob):

    if not dob:
        return None

    today = date.today()

    age = (
        today.year
        - dob.year
        - (
            (today.month, today.day)
            < (dob.month, dob.day)
        )
    )

    return age


def month_name(month):

    if month is None:
        return ""

    if 1 <= month <= 12:
        return calendar.month_name[month]

    return ""


def get_next_month(year, month):

    if month == 12:

        return date(
            year + 1,
            1,
            1,
        )

    return date(
        year,
        month + 1,
        1,
    )


def calculate_student_lifecycle_fee(
    student: Student,
    batch: Batch,
    target_year: int,
    target_month: int,
    all_payments: list[FeePayment],
    today: date,
) -> dict:
    """
    Calculates student fee statistics from their join date up to the target month/year:
    - total_due_months: billing months between join_date and target (year, month)
    - months_paid_count: count of months with full fee paid
    - unpaid_months_count: count of overdue / unpaid months
    - total_pending_amount: cumulative overdue balance
    - fee_status: status for target month (PAID, OVERDUE, DUE TODAY, UNPAID, NOT JOINED)
    - target_paid_amount: amount paid for target month
    - target_pending_amount: balance due for target month
    - target_paid_date, target_payment_method
    """
    monthly_fee = int(
        student.monthly_fee
        if (student.monthly_fee is not None and student.monthly_fee > 0)
        else (batch.monthly_fee or 0)
    )

    join_date = student.join_date or (
        student.created_at.date()
        if getattr(student, "created_at", None)
        else date(target_year, target_month, 1)
    )

    # Group all payments by (fee_year, fee_month)
    paid_map: dict[tuple[int, int], dict] = {}
    lifetime_total_paid = 0

    for p in all_payments:
        amt = int(p.net_payable or 0)
        lifetime_total_paid += amt
        if p.fee_year and p.fee_month:
            k = (p.fee_year, p.fee_month)
            if k not in paid_map:
                paid_map[k] = {"total": 0, "payments": []}
            paid_map[k]["total"] += amt
            paid_map[k]["payments"].append(p)

    target_period = (target_year, target_month)
    join_period = (join_date.year, join_date.month)

    total_due_months = 0
    months_paid_count = 0
    unpaid_months_count = 0
    total_pending_amount = 0

    if join_period <= target_period:
        curr_y = join_date.year
        curr_m = join_date.month

        while (curr_y < target_year) or (curr_y == target_year and curr_m <= target_month):
            total_due_months += 1
            month_info = paid_map.get((curr_y, curr_m), {"total": 0, "payments": []})
            paid_for_m = month_info["total"]

            if (monthly_fee > 0 and paid_for_m >= monthly_fee) or (monthly_fee == 0 and (paid_for_m > 0 or len(month_info["payments"]) > 0)):
                months_paid_count += 1
            else:
                unpaid_months_count += 1
                total_pending_amount += max(0, monthly_fee - paid_for_m)

            if curr_m == 12:
                curr_y += 1
                curr_m = 1
            else:
                curr_m += 1

    # Selected / Target Month Specifics
    target_info = paid_map.get(target_period, {"total": 0, "payments": []})
    target_paid_amt = target_info["total"]
    target_payments = target_info["payments"]

    # Target payment details (latest payment for that month)
    latest_target_payment = target_payments[0] if target_payments else None
    target_paid_date = format_date(latest_target_payment.payment_date) if (latest_target_payment and latest_target_payment.payment_date) else ""
    target_payment_method = str(latest_target_payment.payment_method) if (latest_target_payment and latest_target_payment.payment_method) else ""

    due_date = date(target_year, target_month, 1)

    if join_period > target_period:
        fee_status = "NOT JOINED"
        target_pending_amt = 0
    elif (monthly_fee > 0 and target_paid_amt >= monthly_fee) or (monthly_fee == 0 and (target_paid_amt > 0 or len(target_payments) > 0)):
        fee_status = "PAID"
        target_pending_amt = 0
    else:
        target_pending_amt = max(0, monthly_fee - target_paid_amt)
        if today > due_date:
            fee_status = "OVERDUE"
        elif today == due_date:
            fee_status = "DUE TODAY"
        else:
            fee_status = "UNPAID"

    # Most recent payment ever made by student
    last_payment = all_payments[0] if all_payments else None
    last_paid_amount = int(last_payment.net_payable or 0) if last_payment else 0
    last_paid_month = (
        f"{month_name(last_payment.fee_month)} {last_payment.fee_year or ''}".strip()
        if (last_payment and last_payment.fee_month)
        else "-"
    )
    last_paid_date = (
        format_date(last_payment.payment_date)
        if (last_payment and last_payment.payment_date)
        else "-"
    )

    return {
        "monthly_fee": monthly_fee,
        "fee_status": fee_status,
        "total_due_months": total_due_months,
        "months_paid_count": months_paid_count,
        "unpaid_months_count": unpaid_months_count,
        "total_pending_amount": total_pending_amount,
        "target_paid_amount": target_paid_amt,
        "target_pending_amount": target_pending_amt,
        "target_paid_date": target_paid_date,
        "target_payment_method": target_payment_method,
        "last_paid_amount": last_paid_amount,
        "last_paid_month": last_paid_month,
        "last_paid_date": last_paid_date,
        "lifetime_total_paid": lifetime_total_paid,
    }


# =========================================================
# GENERIC EXCEL EXPORT
# =========================================================


def create_excel(
    sheets: dict,
):

    from openpyxl import Workbook

    from openpyxl.styles import (
        Font,
        Alignment,
    )

    from openpyxl.utils import (
        get_column_letter,
    )

    workbook = Workbook()

    first_sheet = True

    for sheet_name, rows in sheets.items():

        if first_sheet:

            worksheet = workbook.active

            worksheet.title = sheet_name

            first_sheet = False

        else:

            worksheet = (
                workbook.create_sheet(
                    sheet_name
                )
            )

        if not rows:

            worksheet["A1"] = (
                "No data available"
            )

            continue

        headers = list(
            rows[0].keys()
        )

        # -------------------------------------------------
        # HEADERS
        # -------------------------------------------------

        for column_index, header in enumerate(
            headers,
            start=1,
        ):

            cell = worksheet.cell(
                row=1,
                column=column_index,
            )

            cell.value = header

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # -------------------------------------------------
        # DATA
        # -------------------------------------------------

        for row_index, row in enumerate(
            rows,
            start=2,
        ):

            for column_index, header in enumerate(
                headers,
                start=1,
            ):

                worksheet.cell(
                    row=row_index,
                    column=column_index,
                ).value = row.get(
                    header,
                    "",
                )

        # -------------------------------------------------
        # FREEZE
        # -------------------------------------------------

        worksheet.freeze_panes = "A2"

        # -------------------------------------------------
        # WIDTH
        # -------------------------------------------------

        for column_cells in (
            worksheet.columns
        ):

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                length = len(
                    str(
                        cell.value
                        or ""
                    )
                )

                max_length = max(
                    max_length,
                    length,
                )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max(
                    max_length + 2,
                    12,
                ),
                40,
            )

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# =========================================================
# GENERIC CSV EXPORT
# =========================================================


def create_csv(rows):

    output = StringIO()

    if not rows:

        output.write(
            "No data available\n"
        )

    else:

        writer = csv.DictWriter(
            output,
            fieldnames=list(
                rows[0].keys()
            ),
        )

        writer.writeheader()

        writer.writerows(rows)

    output.seek(0)

    return output


# =========================================================
# GENERIC PDF EXPORT
# =========================================================


def create_pdf(
    title,
    rows,
):

    from reportlab.lib import colors

    from reportlab.lib.pagesizes import (
        A4,
        landscape,
    )

    from reportlab.lib.styles import (
        getSampleStyleSheet,
    )

    from reportlab.lib.units import mm

    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=8 * mm,
        leftMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = (
        getSampleStyleSheet()
    )

    story = []

    story.append(
        Paragraph(
            title,
            styles["Title"],
        )
    )

    story.append(
        Spacer(
            1,
            5 * mm,
        )
    )

    if not rows:

        story.append(
            Paragraph(
                "No data available",
                styles["Normal"],
            )
        )

        document.build(story)

        output.seek(0)

        return output

    headers = list(
        rows[0].keys()
    )

    from reportlab.lib.styles import ParagraphStyle

    num_cols = max(len(headers), 1)
    col_width = (281 * mm) / num_cols

    header_cell_style = ParagraphStyle(
        "GenericHeaderCell",
        fontName="Helvetica-Bold",
        fontSize=6.8,
        leading=8.5,
        textColor=colors.white,
        alignment=1,
    )

    cell_style = ParagraphStyle(
        "GenericBodyCell",
        fontName="Helvetica",
        fontSize=6.5,
        leading=8.5,
        textColor=colors.HexColor("#1F2937"),
        alignment=1,
    )

    table_data = [
        [
            Paragraph(str(h), header_cell_style)
            for h in headers
        ]
    ]

    for row in rows:
        table_data.append(
            [
                Paragraph(
                    str(
                        row.get(
                            header,
                            "",
                        )
                    )[:100],
                    cell_style,
                )
                for header in headers
            ]
        )

    table = Table(
        table_data,
        colWidths=[col_width] * num_cols,
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor(
                        "#111827"
                    ),
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white,
                ),
                (
                    "ALIGN",
                    (0, 0),
                    (-1, -1),
                    "CENTER",
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "MIDDLE",
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.3,
                    colors.HexColor("#CBD5E1"),
                ),
                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    2.5,
                ),
                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    2.5,
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    2,
                ),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor(
                            "#F9FAFB"
                        ),
                    ],
                ),
            ]
        )
    )

    story.append(table)

    document.build(story)

    output.seek(0)

    return output


# =========================================================
# STUDENT PDF EXPORT
# =========================================================


def create_student_pdf(
    batch_name: str,
    location: str,
    status_filter: str | None,
    students_data: list[dict],
    is_multi_batch: bool = False,
    batches_data: dict | None = None,
    month_label: str | None = None,
):

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import (
        A4,
        landscape,
    )
    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle,
    )
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "ReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#111827"),
        spaceAfter=0,
    )

    batch_banner_style = ParagraphStyle(
        "BatchBanner",
        fontName="Helvetica-Bold",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#1E3A8A"),
    )

    header_cell_style = ParagraphStyle(
        "HeaderCell",
        fontName="Helvetica-Bold",
        fontSize=5.6,
        leading=7.0,
        textColor=colors.white,
        alignment=1,
    )

    cell_style = ParagraphStyle(
        "BodyCell",
        fontName="Helvetica",
        fontSize=5.3,
        leading=6.8,
        textColor=colors.HexColor("#1F2937"),
        alignment=1,
    )

    cell_style_left = ParagraphStyle(
        "BodyCellLeft",
        fontName="Helvetica",
        fontSize=5.3,
        leading=6.8,
        textColor=colors.HexColor("#1F2937"),
        alignment=0,
    )

    meta_label_style = ParagraphStyle(
        "MetaLabel",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#374151"),
    )

    meta_value_style = ParagraphStyle(
        "MetaValue",
        fontName="Helvetica",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#111827"),
    )

    meta_value_highlight = ParagraphStyle(
        "MetaValueHighlight",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#1D4ED8"),
    )

    story = []

    # 1. Header Title
    story.append(
        Paragraph("Student Report", title_style)
    )
    story.append(Spacer(1, 2.5 * mm))

    # 2. Metadata / Summary Card
    total_students = len(students_data)
    active_count = sum(
        1
        for s in students_data
        if str(s.get("status", "")).upper() == "ACTIVE"
    )
    inactive_count = total_students - active_count
    paid_count = sum(
        1
        for s in students_data
        if str(s.get("fee_status", "")).upper() == "PAID"
    )
    pending_count = total_students - paid_count
    total_pending_amount = sum(
        int(s.get("total_pending_amount", 0))
        for s in students_data
    )

    filter_text = (
        status_filter.upper()
        if status_filter
        else "ALL"
    )
    today_str = date.today().strftime("%d %b %Y")

    summary_data = [
        [
            Paragraph("Batch Name:", meta_label_style),
            Paragraph(batch_name or "All Batches", meta_value_highlight),
            Paragraph("Location:", meta_label_style),
            Paragraph(
                location or ("All Locations" if is_multi_batch else "N/A"),
                meta_value_style,
            ),
            Paragraph("Target Month:", meta_label_style),
            Paragraph(month_label or today_str, meta_value_highlight),
        ],
        [
            Paragraph("Status Filter:", meta_label_style),
            Paragraph(filter_text, meta_value_style),
            Paragraph("Total Students:", meta_label_style),
            Paragraph(
                f"<b>{total_students}</b> (<font color='#047857'>{active_count} Active</font> / <font color='#B91C1C'>{inactive_count} Inactive</font>)",
                meta_value_style,
            ),
            Paragraph("Fee Summary:", meta_label_style),
            Paragraph(
                f"<font color='#047857'><b>{paid_count} Paid</b></font> / <font color='#B91C1C'><b>{pending_count} Pending</b></font> (Due: <b>Rs. {total_pending_amount:,}</b>)",
                meta_value_style,
            ),
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[25 * mm, 65 * mm, 25 * mm, 65 * mm, 32 * mm, 69 * mm],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    story.append(summary_table)
    story.append(Spacer(1, 3.5 * mm))

    if not students_data:
        story.append(
            Paragraph(
                "No students found matching criteria",
                styles["Normal"],
            )
        )
        document.build(story)
        output.seek(0)
        return output

    # 3. Table Column Widths (Total: 281 mm, landscape A4)
    col_widths = [
        7 * mm,     # ID
        25 * mm,    # Student Name
        10 * mm,    # Gender
        7 * mm,     # Age
        14 * mm,    # DOB
        9 * mm,     # Blood
        14 * mm,    # Join Date
        20 * mm,    # Parent Name
        17 * mm,    # Phone
        11 * mm,    # Status
        14 * mm,    # Classes (Att/Cond)
        13 * mm,    # Monthly Fee
        13 * mm,    # Fee Status
        12 * mm,    # Paid Amt
        12 * mm,    # Pending
        14 * mm,    # Paid Date
        11 * mm,    # Method
        14 * mm,    # Months (Paid/Due)
        11 * mm,    # Months Overdue
        14 * mm,    # Total Pending
        14 * mm,    # Last Paid Amt
        15 * mm,    # Last Paid Month
    ]

    headers = [
        "ID",
        "Student Name",
        "Gender",
        "Age",
        "DOB",
        "Blood",
        "Join Date",
        "Parent Name",
        "Phone",
        "Status",
        "Classes<br/>(Att/Cond)",
        "Monthly<br/>Fee",
        "Fee Status",
        "Paid<br/>Amt",
        "Pending",
        "Paid Date",
        "Method",
        "Months<br/>(Paid/Due)",
        "Months<br/>Overdue",
        "Total<br/>Pending",
        "Last Paid<br/>Amt",
        "Last Paid<br/>Month",
    ]

    def build_table_for_students(student_group):
        header_row = [
            Paragraph(h, header_cell_style) for h in headers
        ]
        t_data = [header_row]

        for s in student_group:
            attended = s.get("attended", 0)
            conducted = s.get("conducted", 0)
            classes_str = f"{attended}/{conducted}"

            status_val = str(s.get("status", "ACTIVE")).upper()
            if status_val == "ACTIVE":
                status_para = Paragraph(
                    "<font color='#047857'><b>ACTIVE</b></font>",
                    cell_style,
                )
            else:
                status_para = Paragraph(
                    "<font color='#B91C1C'><b>INACTIVE</b></font>",
                    cell_style,
                )

            fee_status_val = str(s.get("fee_status", "UNPAID")).upper()
            if fee_status_val == "PAID":
                fee_color = "#047857"
            elif fee_status_val in ("OVERDUE", "UNPAID"):
                fee_color = "#B91C1C"
            elif fee_status_val == "DUE TODAY":
                fee_color = "#D97706"
            else:
                fee_color = "#374151"
            fee_status_para = Paragraph(
                f"<font color='{fee_color}'><b>{fee_status_val}</b></font>",
                cell_style,
            )

            months_paid = s.get("months_paid_count", 0)
            total_due_mths = s.get("total_due_months", 0)
            overdue_mths = s.get("unpaid_months_count", 0)
            total_pending = int(s.get("total_pending_amount", 0))
            paid_amount = int(s.get("paid_amount", 0))
            pending_amount = int(s.get("pending_amount", 0))
            last_paid_amount = int(s.get("last_paid_amount", 0))
            last_paid_month_str = str(s.get("last_paid_month", "") or "-")

            paid_due_para = Paragraph(
                f"<b>{months_paid}</b>/{total_due_mths}",
                cell_style,
            )

            overdue_para = Paragraph(
                f"<font color='#B91C1C'><b>{overdue_mths}</b></font>" if overdue_mths > 0 else "<font color='#047857'>0</font>",
                cell_style,
            )

            pending_total_para = Paragraph(
                f"<font color='#B91C1C'><b>{total_pending:,}</b></font>" if total_pending > 0 else "0",
                cell_style,
            )

            month_paid_str = f"{paid_amount:,}" if paid_amount > 0 else "-"
            month_pending_str = f"{pending_amount:,}" if pending_amount > 0 else "0"
            last_paid_amt_str = f"{last_paid_amount:,}" if last_paid_amount > 0 else "-"

            t_data.append(
                [
                    Paragraph(str(s.get("id", "")), cell_style),
                    Paragraph(str(s.get("name", "") or "-"), cell_style_left),
                    Paragraph(str(s.get("gender", "") or "-"), cell_style),
                    Paragraph(str(s.get("age", "") if s.get("age") is not None else "-"), cell_style),
                    Paragraph(str(s.get("dob", "") or "-"), cell_style),
                    Paragraph(str(s.get("blood_group", "") or "-"), cell_style),
                    Paragraph(str(s.get("join_date", "") or "-"), cell_style),
                    Paragraph(str(s.get("parent_name", "") or "-"), cell_style_left),
                    Paragraph(str(s.get("phone", "") or "-"), cell_style),
                    status_para,
                    Paragraph(f"<b>{classes_str}</b>", cell_style),
                    Paragraph(str(s.get("monthly_fee", 0)), cell_style),
                    fee_status_para,
                    Paragraph(month_paid_str, cell_style),
                    Paragraph(month_pending_str, cell_style),
                    Paragraph(str(s.get("paid_date", "") or "-"), cell_style),
                    Paragraph(str(s.get("payment_method", "") or "-"), cell_style),
                    paid_due_para,
                    overdue_para,
                    pending_total_para,
                    Paragraph(last_paid_amt_str, cell_style),
                    Paragraph(last_paid_month_str, cell_style),
                ]
            )

        st = Table(
            t_data,
            colWidths=col_widths,
            repeatRows=1,
        )

        st.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [
                            colors.white,
                            colors.HexColor("#F8FAFC"),
                        ],
                    ),
                ]
            )
        )
        return st

    if is_multi_batch and batches_data:
        for b_name, b_info in batches_data.items():
            b_loc = b_info.get("location", "N/A")
            b_students = b_info.get("students", [])
            banner_text = f"<b>Batch:</b> {b_name} &nbsp;&nbsp;|&nbsp;&nbsp; <b>Location:</b> {b_loc} &nbsp;&nbsp;|&nbsp;&nbsp; <b>Students:</b> {len(b_students)}"
            banner_table = Table(
                [[Paragraph(banner_text, batch_banner_style)]],
                colWidths=[277 * mm],
            )
            banner_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
                        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFDBFE")),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(banner_table)
            story.append(Spacer(1, 1.5 * mm))
            story.append(build_table_for_students(b_students))
            story.append(Spacer(1, 4 * mm))
    else:
        story.append(build_table_for_students(students_data))

    document.build(story)
    output.seek(0)
    return output


# =========================================================
# SINGLE STUDENT PDF EXPORT
# =========================================================


def create_single_student_pdf(
    student_data: dict,
    monthly_ledger: list[dict],
    payment_history: list[dict],
) -> BytesIO:

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle,
    )
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        HRFlowable,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()

    # Styles
    title_style = ParagraphStyle(
        "SingleStudentReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=0,
    )

    subtitle_style = ParagraphStyle(
        "SingleStudentReportSubtitle",
        fontName="Helvetica",
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor("#64748B"),
    )

    section_header_style = ParagraphStyle(
        "SectionHeader",
        fontName="Helvetica-Bold",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#1E293B"),
    )

    label_style = ParagraphStyle(
        "FieldLabel",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#475569"),
    )

    val_style = ParagraphStyle(
        "FieldValue",
        fontName="Helvetica",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#0F172A"),
    )

    val_bold = ParagraphStyle(
        "FieldValueBold",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#0F172A"),
    )

    header_cell_style = ParagraphStyle(
        "THeaderCell",
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=9.0,
        textColor=colors.white,
        alignment=1,
    )

    cell_style = ParagraphStyle(
        "TCell",
        fontName="Helvetica",
        fontSize=6.8,
        leading=8.5,
        textColor=colors.HexColor("#1E293B"),
        alignment=1,
    )

    cell_style_left = ParagraphStyle(
        "TCellLeft",
        fontName="Helvetica",
        fontSize=6.8,
        leading=8.5,
        textColor=colors.HexColor("#1E293B"),
        alignment=0,
    )

    story = []

    # 1. Header Title & Meta
    header_table = Table(
        [
            [
                Paragraph("<b>SKATING ACADEMY</b>", title_style),
                Paragraph(f"<b>Generated:</b> {date.today().strftime('%d %b %Y')}", ParagraphStyle("GenDate", parent=subtitle_style, alignment=2)),
            ],
            [
                Paragraph("Student Profile & Comprehensive Statement", subtitle_style),
                Paragraph(f"<b>Target Month:</b> {student_data.get('month_label', '-')}", ParagraphStyle("TMonth", parent=subtitle_style, alignment=2)),
            ],
        ],
        colWidths=[110 * mm, 80 * mm],
    )
    header_table.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ])
    )
    story.append(header_table)
    story.append(Spacer(1, 2 * mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceAfter=3 * mm))

    # 2. Student Profile Card
    status_str = str(student_data.get("status", "ACTIVE")).upper()
    status_html = "<font color='#047857'><b>ACTIVE</b></font>" if status_str == "ACTIVE" else "<font color='#B91C1C'><b>INACTIVE</b></font>"

    profile_rows = [
        [
            Paragraph("Student Name:", label_style),
            Paragraph(f"<b>{student_data.get('name', '-')}</b>", val_bold),
            Paragraph("Student ID:", label_style),
            Paragraph(str(student_data.get("id", "-")), val_style),
            Paragraph("Status:", label_style),
            Paragraph(status_html, val_style),
        ],
        [
            Paragraph("Gender / Age:", label_style),
            Paragraph(f"{student_data.get('gender', '-')} / {student_data.get('age', '-')} yrs", val_style),
            Paragraph("Date of Birth:", label_style),
            Paragraph(str(student_data.get("dob", "-")), val_style),
            Paragraph("Blood Group:", label_style),
            Paragraph(str(student_data.get("blood_group", "-") or "-"), val_style),
        ],
        [
            Paragraph("Batch Name:", label_style),
            Paragraph(f"<b>{student_data.get('batch_name', '-')}</b>", val_style),
            Paragraph("Location:", label_style),
            Paragraph(str(student_data.get("location", "-")), val_style),
            Paragraph("Join Date:", label_style),
            Paragraph(str(student_data.get("join_date", "-")), val_style),
        ],
        [
            Paragraph("Parent Name:", label_style),
            Paragraph(str(student_data.get("parent_name", "-")), val_style),
            Paragraph("Phone:", label_style),
            Paragraph(str(student_data.get("phone", "-")), val_style),
            Paragraph("Emergency Contact:", label_style),
            Paragraph(str(student_data.get("emergency_contact", "-") or "-"), val_style),
        ],
    ]

    profile_table = Table(
        profile_rows,
        colWidths=[24 * mm, 42 * mm, 24 * mm, 38 * mm, 28 * mm, 34 * mm],
    )
    profile_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
            ("LEFTPADDING", (0, 0), (-1, -1), 3.5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3.5),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    story.append(profile_table)
    story.append(Spacer(1, 3 * mm))

    # 3. KPI / Summary Cards (Attendance & Fee Lifecycle)
    att_pct = student_data.get("attendance_percentage", 0)
    attended = student_data.get("attended", 0)
    conducted = student_data.get("conducted", 0)
    absent = student_data.get("absent", 0)

    m_fee = student_data.get("monthly_fee", 0)
    paid_mths = student_data.get("months_paid_count", 0)
    due_mths = student_data.get("total_due_months", 0)
    overdue_mths = student_data.get("unpaid_months_count", 0)
    tot_pending = student_data.get("total_pending_amount", 0)
    lifetime_paid = student_data.get("lifetime_total_paid", 0)
    curr_fee_status = str(student_data.get("fee_status", "UNPAID")).upper()

    if curr_fee_status == "PAID":
        fee_badge = "<font color='#047857'><b>PAID</b></font>"
    elif curr_fee_status in ("OVERDUE", "UNPAID"):
        fee_badge = "<font color='#B91C1C'><b>OVERDUE</b></font>"
    elif curr_fee_status == "DUE TODAY":
        fee_badge = "<font color='#D97706'><b>DUE TODAY</b></font>"
    else:
        fee_badge = f"<b>{curr_fee_status}</b>"

    kpi_rows = [
        [
            Paragraph("<b>ATTENDANCE SUMMARY</b>", ParagraphStyle("KPIHead1", parent=label_style, textColor=colors.HexColor("#1E3A8A"))),
            Paragraph("<b>FEE & PAYMENT LIFECYCLE</b>", ParagraphStyle("KPIHead2", parent=label_style, textColor=colors.HexColor("#1E3A8A"))),
        ],
        [
            Paragraph(
                f"• <b>Attendance Rate:</b> {att_pct}%<br/>"
                f"• <b>Classes Conducted:</b> {conducted}<br/>"
                f"• <b>Attended:</b> <font color='#047857'><b>{attended}</b></font> &nbsp;|&nbsp; <b>Absent:</b> <font color='#B91C1C'><b>{absent}</b></font>",
                val_style,
            ),
            Paragraph(
                f"• <b>Monthly Fee:</b> Rs. {m_fee:,} &nbsp;|&nbsp; <b>Target Month Status:</b> {fee_badge}<br/>"
                f"• <b>Billing Months:</b> <b>{paid_mths}</b> / {due_mths} Paid &nbsp;|&nbsp; <b>Overdue:</b> <font color='#B91C1C'><b>{overdue_mths} Months</b></font><br/>"
                f"• <b>Lifetime Total Paid:</b> <font color='#047857'><b>Rs. {lifetime_paid:,}</b></font> &nbsp;|&nbsp; <b>Total Pending:</b> <font color='#B91C1C'><b>Rs. {tot_pending:,}</b></font>",
                val_style,
            ),
        ],
    ]

    kpi_table = Table(
        kpi_rows,
        colWidths=[80 * mm, 110 * mm],
    )
    kpi_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EFF6FF")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#BFDBFE")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#DBEAFE")),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ])
    )
    story.append(kpi_table)
    story.append(Spacer(1, 4 * mm))

    # 4. Section: Month-by-Month Fee Statement Table
    story.append(Paragraph("<b>1. Month-by-Month Fee Statement</b>", section_header_style))
    story.append(Spacer(1, 1.5 * mm))

    ledger_headers = [
        "Month & Year",
        "Fee Due",
        "Paid Amt",
        "Pending",
        "Fee Status",
        "Paid Date",
        "Method",
    ]
    ledger_col_widths = [
        32 * mm,
        22 * mm,
        22 * mm,
        25 * mm,
        25 * mm,
        32 * mm,
        32 * mm,
    ]

    ledger_data = [
        [Paragraph(h, header_cell_style) for h in ledger_headers]
    ]

    for row in monthly_ledger:
        st_val = str(row.get("Status", "UNPAID")).upper()
        if st_val == "PAID":
            st_color = "#047857"
        elif st_val in ("OVERDUE", "UNPAID"):
            st_color = "#B91C1C"
        elif st_val == "DUE TODAY":
            st_color = "#D97706"
        else:
            st_color = "#475569"

        st_para = Paragraph(f"<font color='{st_color}'><b>{st_val}</b></font>", cell_style)
        due_amt = int(row.get("Fee Amount", 0))
        paid_amt = int(row.get("Paid Amount", 0))
        pending_amt = int(row.get("Pending Amount", 0))

        paid_str = f"Rs. {paid_amt:,}" if paid_amt > 0 else "-"
        pending_str = f"<font color='#B91C1C'><b>Rs. {pending_amt:,}</b></font>" if pending_amt > 0 else "Rs. 0"

        ledger_data.append([
            Paragraph(str(row.get("Month & Year", "-")), cell_style_left),
            Paragraph(f"Rs. {due_amt:,}", cell_style),
            Paragraph(paid_str, cell_style),
            Paragraph(pending_str, cell_style),
            st_para,
            Paragraph(str(row.get("Paid Date", "-")), cell_style),
            Paragraph(str(row.get("Payment Method", "-")), cell_style),
        ])

    ledger_table = Table(
        ledger_data,
        colWidths=ledger_col_widths,
        repeatRows=1,
    )
    ledger_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            (
                "ROWBACKGROUNDS",
                (0, 1),
                (-1, -1),
                [
                    colors.white,
                    colors.HexColor("#F8FAFC"),
                ],
            ),
        ])
    )
    story.append(ledger_table)
    story.append(Spacer(1, 4 * mm))

    # 5. Section: Payment Transaction History
    story.append(Paragraph("<b>2. Payment Receipts & Transaction History</b>", section_header_style))
    story.append(Spacer(1, 1.5 * mm))

    if payment_history:
        p_headers = [
            "Receipt ID",
            "Payment Date",
            "Fee Month",
            "Fee Year",
            "Amount Paid",
            "Method",
            "Status",
        ]
        p_col_widths = [
            22 * mm,
            36 * mm,
            30 * mm,
            22 * mm,
            30 * mm,
            28 * mm,
            22 * mm,
        ]

        p_data = [
            [Paragraph(h, header_cell_style) for h in p_headers]
        ]

        for p_item in payment_history:
            amt = int(p_item.get("Amount", 0))
            p_data.append([
                Paragraph(str(p_item.get("Payment ID", "-")), cell_style),
                Paragraph(str(p_item.get("Payment Date", "-")), cell_style),
                Paragraph(str(p_item.get("Fee Month", "-")), cell_style),
                Paragraph(str(p_item.get("Fee Year", "-")), cell_style),
                Paragraph(f"<b>Rs. {amt:,}</b>", cell_style),
                Paragraph(str(p_item.get("Payment Method", "-")), cell_style),
                Paragraph("<font color='#047857'><b>PAID</b></font>", cell_style),
            ])

        p_table = Table(
            p_data,
            colWidths=p_col_widths,
            repeatRows=1,
        )
        p_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#F8FAFC"),
                    ],
                ),
            ])
        )
        story.append(p_table)
    else:
        story.append(Paragraph("<i>No payment transactions recorded yet.</i>", subtitle_style))

    document.build(story)
    output.seek(0)
    return output


# =========================================================
# BATCH PDF EXPORT
# =========================================================


def create_batch_pdf(
    batches_data: list[dict],
    batch_filter_name: str | None = None,
    month_label: str | None = None,
) -> BytesIO:

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import (
        A4,
        landscape,
    )
    from reportlab.lib.styles import (
        getSampleStyleSheet,
        ParagraphStyle,
    )
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        "BatchReportTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=18,
        textColor=colors.HexColor("#111827"),
        spaceAfter=0,
    )

    header_cell_style = ParagraphStyle(
        "BatchHeaderCell",
        fontName="Helvetica-Bold",
        fontSize=6.8,
        leading=8.5,
        textColor=colors.white,
        alignment=1,
    )

    cell_style = ParagraphStyle(
        "BatchBodyCell",
        fontName="Helvetica",
        fontSize=6.5,
        leading=8.5,
        textColor=colors.HexColor("#1F2937"),
        alignment=1,
    )

    cell_style_left = ParagraphStyle(
        "BatchBodyCellLeft",
        fontName="Helvetica",
        fontSize=6.5,
        leading=8.5,
        textColor=colors.HexColor("#1F2937"),
        alignment=0,
    )

    meta_label_style = ParagraphStyle(
        "BatchMetaLabel",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#374151"),
    )

    meta_value_style = ParagraphStyle(
        "BatchMetaValue",
        fontName="Helvetica",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#111827"),
    )

    meta_value_highlight = ParagraphStyle(
        "BatchMetaValueHighlight",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        leading=9.5,
        textColor=colors.HexColor("#1D4ED8"),
    )

    story = []

    # 1. Title
    story.append(Paragraph("Batch Report", title_style))
    story.append(Spacer(1, 2.5 * mm))

    # 2. Metadata / Summary Card
    total_batches = len(batches_data)
    total_students = sum(b.get("students", 0) for b in batches_data)
    total_revenue = sum(b.get("revenue", 0) for b in batches_data)
    total_pending = sum(b.get("pending_fee", 0) for b in batches_data)
    today_str = date.today().strftime("%d %b %Y")

    filter_text = batch_filter_name if batch_filter_name else "ALL BATCHES"

    summary_data = [
        [
            Paragraph("Batch Filter:", meta_label_style),
            Paragraph(filter_text, meta_value_highlight),
            Paragraph("Total Batches:", meta_label_style),
            Paragraph(str(total_batches), meta_value_style),
            Paragraph("Target Month:", meta_label_style),
            Paragraph(month_label or today_str, meta_value_highlight),
        ],
        [
            Paragraph("Total Students:", meta_label_style),
            Paragraph(str(total_students), meta_value_style),
            Paragraph("Month Revenue:", meta_label_style),
            Paragraph(f"<font color='#047857'><b>Rs. {total_revenue:,}</b></font>", meta_value_style),
            Paragraph("Total Pending Fees:", meta_label_style),
            Paragraph(f"<font color='#B91C1C'><b>Rs. {total_pending:,}</b></font>", meta_value_style),
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[25 * mm, 65 * mm, 30 * mm, 60 * mm, 32 * mm, 69 * mm],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )

    story.append(summary_table)
    story.append(Spacer(1, 3.5 * mm))

    if not batches_data:
        story.append(Paragraph("No batches found matching criteria", styles["Normal"]))
        document.build(story)
        output.seek(0)
        return output

    # 3. Table Column Widths (Total: 281 mm)
    col_widths = [
        10 * mm,  # ID
        30 * mm,  # Batch Name
        25 * mm,  # Location
        17 * mm,  # Level
        18 * mm,  # Class Type
        37 * mm,  # Training Days
        19 * mm,  # Monthly Fee
        15 * mm,  # Students
        17 * mm,  # Conducted
        15 * mm,  # Present
        15 * mm,  # Absent
        18 * mm,  # Attendance %
        22 * mm,  # This Month Rev
        23 * mm,  # Pending Fees
    ]

    headers = [
        "ID",
        "Batch Name",
        "Location",
        "Level",
        "Class Type",
        "Training Days",
        "Monthly<br/>Fee",
        "Students",
        "Classes<br/>Conducted",
        "Present",
        "Absent",
        "Attendance<br/>%",
        "Month<br/>Revenue",
        "Pending<br/>Fees",
    ]

    header_row = [
        Paragraph(h, header_cell_style) for h in headers
    ]
    table_data = [header_row]

    for b in batches_data:
        days_val = b.get("training_days", "")
        if isinstance(days_val, list):
            days_str = ", ".join(days_val)
        else:
            days_str = str(days_val or "-")

        for full_day, short_day in [
            ("Monday", "Mon"), ("Tuesday", "Tue"), ("Wednesday", "Wed"),
            ("Thursday", "Thu"), ("Friday", "Fri"), ("Saturday", "Sat"), ("Sunday", "Sun")
        ]:
            days_str = days_str.replace(full_day, short_day)

        att_pct = b.get("attendance_percentage", 0)

        table_data.append(
            [
                Paragraph(str(b.get("id", "")), cell_style),
                Paragraph(str(b.get("batch_name", "") or "-"), cell_style_left),
                Paragraph(str(b.get("location", "") or "-"), cell_style_left),
                Paragraph(str(b.get("level", "") or "-"), cell_style),
                Paragraph(str(b.get("class_type", "") or "-"), cell_style),
                Paragraph(days_str, cell_style_left),
                Paragraph(str(b.get("monthly_fee", 0)), cell_style),
                Paragraph(str(b.get("students", 0)), cell_style),
                Paragraph(str(b.get("conducted", 0)), cell_style),
                Paragraph(str(b.get("present", 0)), cell_style),
                Paragraph(str(b.get("absent", 0)), cell_style),
                Paragraph(f"{att_pct:.1f}%", cell_style),
                Paragraph(f"Rs. {b.get('revenue', 0):,}", cell_style),
                Paragraph(
                    f"<font color='#B91C1C'><b>Rs. {b.get('pending_fee', 0):,}</b></font>"
                    if b.get("pending_fee", 0) > 0
                    else "Rs. 0",
                    cell_style,
                ),
            ]
        )

    batch_table = Table(
        table_data,
        colWidths=col_widths,
        repeatRows=1,
    )

    batch_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [
                        colors.white,
                        colors.HexColor("#F8FAFC"),
                    ],
                ),
            ]
        )
    )

    story.append(batch_table)
    document.build(story)
    output.seek(0)
    return output


# =========================================================
# FEE OVERVIEW PDF EXPORT
# =========================================================


def create_fee_overview_pdf(
    month_label: str,
    kpi_data: dict,
    payment_methods_data: list[dict],
    batches_data: list[dict],
    students_data: list[dict],
    batch_filter_name: str | None = None,
    fee_status_filter: str | None = None,
    payment_method_filter: str | None = None,
) -> BytesIO:

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        HRFlowable,
    )

    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = getSampleStyleSheet()

    # Typography Styles
    title_style = ParagraphStyle(
        "FeeOverviewTitle",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        textColor=colors.HexColor("#1E3A8A"),
        spaceAfter=0,
    )

    subtitle_style = ParagraphStyle(
        "FeeOverviewSubtitle",
        fontName="Helvetica",
        fontSize=8,
        leading=10.5,
        textColor=colors.HexColor("#64748B"),
    )

    section_header_style = ParagraphStyle(
        "FeeSectionHeader",
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=11.5,
        textColor=colors.HexColor("#1E293B"),
    )

    meta_label_style = ParagraphStyle(
        "FeeMetaLabel",
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=9.2,
        textColor=colors.HexColor("#374151"),
    )

    meta_value_style = ParagraphStyle(
        "FeeMetaValue",
        fontName="Helvetica",
        fontSize=7.2,
        leading=9.2,
        textColor=colors.HexColor("#111827"),
    )

    meta_value_highlight = ParagraphStyle(
        "FeeMetaValueHighlight",
        fontName="Helvetica-Bold",
        fontSize=7.2,
        leading=9.2,
        textColor=colors.HexColor("#1D4ED8"),
    )

    kpi_section_header = ParagraphStyle(
        "FeeKPISectionHeader",
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.8,
        textColor=colors.HexColor("#1E3A8A"),
        alignment=0,
    )

    header_cell_style = ParagraphStyle(
        "FeeTHeader",
        fontName="Helvetica-Bold",
        fontSize=6.2,
        leading=7.8,
        textColor=colors.white,
        alignment=1,
    )

    cell_style = ParagraphStyle(
        "FeeTCell",
        fontName="Helvetica",
        fontSize=5.5,
        leading=7.0,
        textColor=colors.HexColor("#1E293B"),
        alignment=1,
    )

    cell_style_left = ParagraphStyle(
        "FeeTCellLeft",
        fontName="Helvetica",
        fontSize=5.5,
        leading=7.0,
        textColor=colors.HexColor("#1E293B"),
        alignment=0,
    )

    story = []

    # 1. Header Title Banner
    filters_applied = []
    if batch_filter_name:
        filters_applied.append(f"Batch: {batch_filter_name}")
    if fee_status_filter:
        filters_applied.append(f"Fee Status: {fee_status_filter.upper()}")
    if payment_method_filter:
        filters_applied.append(f"Method: {payment_method_filter.upper()}")
    filter_str = ", ".join(filters_applied) if filters_applied else "All Data"

    header_table = Table(
        [
            [
                Paragraph("<b>SKATING ACADEMY - FEE OVERVIEW & FINANCIAL REPORT</b>", title_style),
                Paragraph(f"<b>Generated:</b> {date.today().strftime('%d %b %Y')}", ParagraphStyle("GenD", parent=subtitle_style, alignment=2)),
            ],
            [
                Paragraph(f"<b>Target Month:</b> {month_label} &nbsp;|&nbsp; <b>Filters:</b> {filter_str}", subtitle_style),
                Paragraph("Executive Financial Statement", ParagraphStyle("SubR", parent=subtitle_style, alignment=2)),
            ],
        ],
        colWidths=[180 * mm, 101 * mm],
    )
    header_table.setStyle(
        TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ])
    )
    story.append(header_table)
    story.append(Spacer(1, 1.5 * mm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#CBD5E1"), spaceAfter=2.5 * mm))

    # 2. Executive Financial Summary & KPI Card (Full Width: 281 mm)
    exp_rev = kpi_data.get("total_expected_revenue", 0)
    coll_rev = kpi_data.get("total_collected_revenue", 0)
    pend_dues = kpi_data.get("total_pending_dues", 0)
    rec_rate = kpi_data.get("collection_rate", 0)

    tot_stu = kpi_data.get("total_students_count", 0)
    paid_stu = kpi_data.get("paid_students_count", 0)
    overdue_stu = kpi_data.get("overdue_students_count", 0)
    due_today_stu = kpi_data.get("due_today_students_count", 0)
    unpaid_stu = kpi_data.get("unpaid_students_count", 0)

    summary_card_data = [
        # Row 0: Metadata Row (Image 1 Style)
        [
            Paragraph("Batch Filter:", meta_label_style),
            Paragraph(batch_filter_name or "ALL BATCHES", meta_value_highlight),
            Paragraph("Total Batches:", meta_label_style),
            Paragraph(str(len(batches_data)), meta_value_style),
            Paragraph("Target Month:", meta_label_style),
            Paragraph(month_label, meta_value_highlight),
        ],
        # Row 1: KPI Section Headers
        [
            Paragraph("<b>REVENUE & COLLECTION HEALTH</b>", kpi_section_header),
            "",
            "",
            Paragraph("<b>STUDENT FEE STATUS BREAKDOWN</b>", kpi_section_header),
            "",
            "",
        ],
        # Row 2: Metrics Line 1
        [
            Paragraph("• Expected Revenue:", meta_label_style),
            Paragraph(f"<b>Rs. {exp_rev:,}</b>", meta_value_style),
            "",
            Paragraph("• Total Active Students:", meta_label_style),
            Paragraph(f"<b>{tot_stu}</b>", meta_value_style),
            "",
        ],
        # Row 3: Metrics Line 2
        [
            Paragraph("• Collected Revenue:", meta_label_style),
            Paragraph(f"<font color='#047857'><b>Rs. {coll_rev:,}</b></font>", meta_value_style),
            "",
            Paragraph("• Paid Students:", meta_label_style),
            Paragraph(f"<font color='#047857'><b>{paid_stu}</b></font> ({((paid_stu/tot_stu)*100 if tot_stu else 0):.1f}%)", meta_value_style),
            "",
        ],
        # Row 4: Metrics Line 3
        [
            Paragraph("• Pending Dues:", meta_label_style),
            Paragraph(f"<font color='#B91C1C'><b>Rs. {pend_dues:,}</b></font>", meta_value_style),
            "",
            Paragraph("• Overdue Students:", meta_label_style),
            Paragraph(f"<font color='#B91C1C'><b>{overdue_stu}</b></font>", meta_value_style),
            "",
        ],
        # Row 5: Metrics Line 4
        [
            Paragraph("• Collection Recovery Rate:", meta_label_style),
            Paragraph(f"<b>{rec_rate:.1f}%</b>", meta_value_highlight),
            "",
            Paragraph("• Due Today / Unpaid:", meta_label_style),
            Paragraph(f"<font color='#D97706'><b>{due_today_stu + unpaid_stu}</b></font>", meta_value_style),
            "",
        ],
    ]

    summary_table = Table(
        summary_card_data,
        colWidths=[42 * mm, 52 * mm, 46.5 * mm, 42 * mm, 52 * mm, 46.5 * mm],
    )
    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 2.2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2.2),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Spans for Section Header Row
            ("SPAN", (0, 1), (2, 1)),
            ("SPAN", (3, 1), (5, 1)),
            ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#EFF6FF")),
            # Spans for Value Columns
            ("SPAN", (1, 2), (2, 2)),
            ("SPAN", (4, 2), (5, 2)),
            ("SPAN", (1, 3), (2, 3)),
            ("SPAN", (4, 3), (5, 3)),
            ("SPAN", (1, 4), (2, 4)),
            ("SPAN", (4, 4), (5, 4)),
            ("SPAN", (1, 5), (2, 5)),
            ("SPAN", (4, 5), (5, 5)),
        ])
    )
    story.append(summary_table)
    story.append(Spacer(1, 3 * mm))

    # 3. Payment Mode & Channels Summary Table (Full Width: 281 mm)
    if payment_methods_data:
        story.append(Paragraph("<b>1. Payment Mode & Collection Channel Split</b>", section_header_style))
        story.append(Spacer(1, 1.2 * mm))

        pm_headers = ["Payment Mode", "Transactions Count", "Amount Collected (Rs.)", "% Share of Total"]
        pm_col_widths = [80 * mm, 60 * mm, 75 * mm, 66 * mm]

        pm_data = [[Paragraph(h, header_cell_style) for h in pm_headers]]
        for pm in payment_methods_data:
            pm_data.append([
                Paragraph(str(pm.get("method", "-")), cell_style_left),
                Paragraph(str(pm.get("count", 0)), cell_style),
                Paragraph(f"Rs. {int(pm.get('amount', 0)):,}", cell_style),
                Paragraph(f"{pm.get('share_pct', 0):.1f}%", cell_style),
            ])

        pm_table = Table(pm_data, colWidths=pm_col_widths, repeatRows=1)
        pm_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#334155")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ])
        )
        story.append(pm_table)
        story.append(Spacer(1, 3.5 * mm))

    # 4. Batch-Wise Performance Table (Full Width: 281 mm)
    if batches_data:
        story.append(Paragraph("<b>2. Batch-Wise Financial & Collection Performance</b>", section_header_style))
        story.append(Spacer(1, 1.2 * mm))

        b_headers = [
            "Batch Name",
            "Location",
            "Students",
            "Monthly Fee",
            "Expected (Rs.)",
            "Collected (Rs.)",
            "Pending (Rs.)",
            "Recovery %",
            "Paid / Total",
        ]
        b_col_widths = [
            46 * mm,
            38 * mm,
            18 * mm,
            24 * mm,
            38 * mm,
            38 * mm,
            38 * mm,
            21 * mm,
            20 * mm,
        ]

        b_data = [[Paragraph(h, header_cell_style) for h in b_headers]]
        for b in batches_data:
            rec_pct = b.get("recovery_pct", 0)
            b_data.append([
                Paragraph(str(b.get("batch_name", "-")), cell_style_left),
                Paragraph(str(b.get("location", "-")), cell_style_left),
                Paragraph(str(b.get("students", 0)), cell_style),
                Paragraph(f"Rs. {b.get('monthly_fee', 0):,}", cell_style),
                Paragraph(f"Rs. {b.get('expected', 0):,}", cell_style),
                Paragraph(f"<font color='#047857'><b>Rs. {b.get('collected', 0):,}</b></font>", cell_style),
                Paragraph(f"<font color='#B91C1C'><b>Rs. {b.get('pending', 0):,}</b></font>" if b.get('pending', 0) > 0 else "Rs. 0", cell_style),
                Paragraph(f"<b>{rec_pct:.1f}%</b>", cell_style),
                Paragraph(str(b.get("paid_vs_total", "-")), cell_style),
            ])

        b_table = Table(b_data, colWidths=b_col_widths, repeatRows=1)
        b_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ])
        )
        story.append(b_table)
        story.append(Spacer(1, 3.5 * mm))

    # 5. Student Fee Dues & Status Table (Full Width: 281 mm)
    if students_data:
        story.append(Paragraph("<b>3. Student Fee Status & Dues Breakdown</b>", section_header_style))
        story.append(Spacer(1, 1.2 * mm))

        s_headers = [
            "ID",
            "Student Name",
            "Batch",
            "Phone",
            "Monthly<br/>Fee",
            "Fee Status",
            "Paid<br/>Amt",
            "Pending",
            "Paid Date",
            "Method",
            "Months<br/>(Paid/Due)",
            "Months<br/>Overdue",
            "Total<br/>Pending",
            "Last Paid<br/>Amt",
            "Last Paid<br/>Month",
        ]
        s_col_widths = [
            8 * mm,
            30 * mm,
            26 * mm,
            22 * mm,
            16 * mm,
            17 * mm,
            16 * mm,
            16 * mm,
            22 * mm,
            16 * mm,
            20 * mm,
            16 * mm,
            22 * mm,
            17 * mm,
            23 * mm,
        ]

        s_data = [[Paragraph(h, header_cell_style) for h in s_headers]]
        for s in students_data:
            f_status = str(s.get("fee_status", "UNPAID")).upper()
            if f_status == "PAID":
                st_color = "#047857"
            elif f_status in ("OVERDUE", "UNPAID"):
                st_color = "#B91C1C"
            elif f_status == "DUE TODAY":
                st_color = "#D97706"
            else:
                st_color = "#475569"

            st_para = Paragraph(f"<font color='{st_color}'><b>{f_status}</b></font>", cell_style)
            paid_due_str = f"<b>{s.get('months_paid_count', 0)}</b>/{s.get('total_due_months', 0)}"
            overdue_m = s.get("unpaid_months_count", 0)
            overdue_para = Paragraph(f"<font color='#B91C1C'><b>{overdue_m}</b></font>" if overdue_m > 0 else "<font color='#047857'>0</font>", cell_style)

            tot_p = int(s.get("total_pending_amount", 0))
            tot_p_para = Paragraph(f"<font color='#B91C1C'><b>Rs. {tot_p:,}</b></font>" if tot_p > 0 else "Rs. 0", cell_style)

            m_paid = int(s.get("paid_amount", 0))
            m_pend = int(s.get("pending_amount", 0))
            lp_amt = int(s.get("last_paid_amount", 0))

            s_data.append([
                Paragraph(str(s.get("id", "")), cell_style),
                Paragraph(str(s.get("name", "") or "-"), cell_style_left),
                Paragraph(str(s.get("batch_name", "") or "-"), cell_style_left),
                Paragraph(str(s.get("phone", "") or "-"), cell_style),
                Paragraph(str(s.get("monthly_fee", 0)), cell_style),
                st_para,
                Paragraph(f"Rs. {m_paid:,}" if m_paid > 0 else "-", cell_style),
                Paragraph(f"Rs. {m_pend:,}" if m_pend > 0 else "0", cell_style),
                Paragraph(str(s.get("paid_date", "") or "-"), cell_style),
                Paragraph(str(s.get("payment_method", "") or "-"), cell_style),
                Paragraph(paid_due_str, cell_style),
                overdue_para,
                tot_p_para,
                Paragraph(f"Rs. {lp_amt:,}" if lp_amt > 0 else "-", cell_style),
                Paragraph(str(s.get("last_paid_month", "") or "-"), cell_style),
            ])

        s_table = Table(s_data, colWidths=s_col_widths, repeatRows=1)
        s_table.setStyle(
            TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, 0), (-1, -1), 1.8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 1.8),
                ("LEFTPADDING", (0, 0), (-1, -1), 1.5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 1.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ])
        )
        story.append(s_table)

    document.build(story)
    output.seek(0)
    return output


# =========================================================
# RESPONSE FILE
# =========================================================


def export_response(
    data,
    format,
    filename,
    title,
):

    if format == "xlsx":

        output = create_excel(
            data
        )

        return StreamingResponse(
            output,
            media_type=(
                "application/"
                "vnd.openxmlformats-"
                "officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition":
                    f'attachment; filename="{filename}.xlsx"'
            },
        )

    if format == "csv":

        # First sheet only
        first_rows = next(
            iter(data.values())
        )

        output = create_csv(
            first_rows
        )

        return StreamingResponse(
            iter([
                output.getvalue()
            ]),
            media_type="text/csv",
            headers={
                "Content-Disposition":
                    f'attachment; filename="{filename}.csv"'
            },
        )

    if format == "pdf":

        first_rows = next(
            iter(data.values())
        )

        output = create_pdf(
            title,
            first_rows,
        )

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                    f'attachment; filename="{filename}.pdf"'
            },
        )

    raise HTTPException(
        status_code=400,
        detail="Unsupported export format",
    )


# =========================================================
# 1. ATTENDANCE EXPORT
# =========================================================


@router.get(
    "/attendance/export"
)
def export_attendance_report(

    from_date: date | None = Query(
        default=None
    ),

    to_date: date | None = Query(
        default=None
    ),

    batch_id: int | None = Query(
        default=None,
        gt=0,
    ),

    format: str = Query(
        default="xlsx",
        pattern="^(xlsx|pdf|csv)$",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):

    today = date.today()

    start_date = (
        from_date
        or today.replace(day=1)
    )

    end_date = (
        to_date
        or today
    )

    if start_date > end_date:

        raise HTTPException(
            status_code=400,
            detail=(
                "from_date cannot be "
                "greater than to_date"
            ),
        )

    # -----------------------------------------------------
    # SESSIONS
    # -----------------------------------------------------

    query = (
        db.query(SessionModel)
        .filter(
            SessionModel.status
            == "COMPLETED",

            SessionModel.session_date
            >= start_date,

            SessionModel.session_date
            <= end_date,
        )
    )

    if batch_id:

        query = query.filter(
            SessionModel.batch_id
            == batch_id
        )

    sessions = (
        query
        .order_by(
            SessionModel.session_date.asc()
        )
        .all()
    )

    student_rows = []
    session_rows = []

    # -----------------------------------------------------
    # SESSION REPORT
    # -----------------------------------------------------

    for session in sessions:

        batch = (
            db.query(Batch)
            .filter(
                Batch.id
                == session.batch_id
            )
            .first()
        )

        expected = (
            db.query(
                func.count(Student.id)
            )
            .filter(
                Student.batch_id
                == session.batch_id,

                Student.is_active.is_(True),
            )
            .scalar()
            or 0
        )

        present = (
            db.query(
                func.count(
                    Attendance.id
                )
            )
            .filter(
                Attendance.session_id
                == session.id,

                Attendance.status
                == "Present",
            )
            .scalar()
            or 0
        )

        absent = max(
            expected - present,
            0,
        )

        percentage = (
            round(
                (
                    present
                    / expected
                )
                * 100,
                2,
            )
            if expected
            else 0
        )

        session_rows.append(
            {
                "Session ID":
                    session.id,

                "Date":
                    format_date(
                        session.session_date
                    ),

                "Batch":
                    batch.batch_name
                    if batch
                    else "",

                "Students":
                    expected,

                "Present":
                    present,

                "Absent":
                    absent,

                "Attendance %":
                    percentage,
            }
        )

        # -------------------------------------------------
        # STUDENT-WISE
        # -------------------------------------------------

        batch_students = (
            db.query(Student)
            .filter(
                Student.batch_id
                == session.batch_id,

                Student.is_active.is_(True),
            )
            .all()
        )

        for student in batch_students:

            attendance = (
                db.query(Attendance)
                .filter(
                    Attendance.session_id
                    == session.id,

                    Attendance.student_id
                    == student.id,
                )
                .first()
            )

            status = (
                attendance.status
                if attendance
                else "Absent"
            )

            student_rows.append(
                {
                    "Session Date":
                        format_date(
                            session.session_date
                        ),

                    "Session ID":
                        session.id,

                    "Student ID":
                        student.id,

                    "Student Name":
                        student.full_name,

                    "Batch Name":
                        batch.batch_name
                        if batch
                        else "",

                    "Attendance":
                        status,
                }
            )

    return export_response(
        {
            "Attendance":
                student_rows,

            "Session Summary":
                session_rows,
        },
        format,
        "attendance-report",
        "Attendance Report",
    )


# =========================================================
# 2. REVENUE EXPORT
# =========================================================


@router.get(
    "/revenue/export"
)
def export_revenue_report(

    from_date: date | None = Query(
        default=None
    ),

    to_date: date | None = Query(
        default=None
    ),

    batch_id: int | None = Query(
        default=None,
        gt=0,
    ),

    payment_method: str | None = Query(
        default=None
    ),

    format: str = Query(
        default="xlsx",
        pattern="^(xlsx|pdf|csv)$",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):

    query = (
        db.query(
            FeePayment,
            Student,
            Batch,
        )
        .join(
            Student,
            FeePayment.student_id
            == Student.id,
        )
        .join(
            Batch,
            Student.batch_id
            == Batch.id,
        )
        .filter(
            Student.is_active.is_(True),
            Batch.is_active.is_(True),
        )
    )

    if from_date:

        query = query.filter(
            FeePayment.payment_date
            >= datetime.combine(
                from_date,
                time.min,
            )
        )

    if to_date:

        query = query.filter(
            FeePayment.payment_date
            < datetime.combine(
                to_date,
                time.max,
            )
        )

    if batch_id:

        query = query.filter(
            Student.batch_id
            == batch_id
        )

    if payment_method:

        query = query.filter(
            FeePayment.payment_method
            == payment_method.upper()
        )

    payments = (
        query
        .order_by(
            FeePayment.payment_date.desc()
        )
        .all()
    )

    rows = []

    for payment, student, batch in payments:

        rows.append(
            {
                "Payment ID":
                    payment.id,

                "Student ID":
                    student.id,

                "Student Name":
                    student.full_name,

                "Batch Name":
                    batch.batch_name,

                "Location":
                    batch.location
                    or "",

                "Fee Month":
                    month_name(
                        payment.fee_month
                    ),

                "Fee Year":
                    payment.fee_year,

                "Amount":
                    int(
                        payment.net_payable
                        or 0
                    ),

                "Payment Method":
                    str(
                        payment.payment_method
                    ),

                "Payment Date":
                    format_date(
                        payment.payment_date
                    ),

                "Status":
                    "PAID",
            }
        )

    return export_response(
        {
            "Revenue":
                rows
        },
        format,
        "revenue-report",
        "Revenue Report",
    )


# =========================================================
# 3. STUDENT EXPORT
# =========================================================


@router.get(
    "/students/export"
)
def export_student_report(

    batch_id: int | None = Query(
        default=None,
        gt=0,
        description="Filter by batch ID",
    ),

    month: int | None = Query(
        default=None,
        ge=1,
        le=12,
        description="Target fee month (1-12, defaults to current month)",
    ),

    year: int | None = Query(
        default=None,
        description="Target fee year (defaults to current year)",
    ),

    status: str | None = Query(
        default=None,
        description="Status filter: active | inactive | paid | overdue | due_today | unpaid | all",
    ),

    fee_status: str | None = Query(
        default=None,
        description="Explicit fee status filter: paid | overdue | due_today | unpaid | all",
    ),

    format: str = Query(
        default="xlsx",
        pattern="^(xlsx|pdf|csv)$",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year
    target_month_label = f"{month_name(target_month)} {target_year}"

    # Determine student active/inactive filter
    effective_status = (status or "").strip().lower()
    effective_fee_status = (fee_status or "").strip().lower()

    if effective_status in ("active", "inactive"):
        active_filter = effective_status
    else:
        active_filter = None

    # Determine fee status filter
    if effective_fee_status in ("paid", "overdue", "due_today", "due today", "unpaid"):
        target_fee_filter = effective_fee_status.replace(" ", "_")
    elif effective_status in ("paid", "overdue", "due_today", "due today", "unpaid"):
        target_fee_filter = effective_status.replace(" ", "_")
    else:
        target_fee_filter = None

    query = (
        db.query(
            Student,
            Batch,
        )
        .join(
            Batch,
            Student.batch_id
            == Batch.id,
        )
    )

    if batch_id:

        query = query.filter(
            Student.batch_id
            == batch_id
        )

    if active_filter == "active":

        query = query.filter(
            Student.is_active.is_(True)
        )

    elif active_filter == "inactive":

        query = query.filter(
            Student.is_active.is_(False)
        )

    students = (
        query
        .order_by(
            Student.id.desc()
        )
        .all()
    )

    rows = []
    payment_rows = []
    pdf_student_items = []

    for student, batch in students:

        # -------------------------------------------------
        # COMPLETED CLASSES
        # -------------------------------------------------

        session_query = (
            db.query(SessionModel)
            .filter(
                SessionModel.batch_id
                == student.batch_id,

                SessionModel.status
                == "COMPLETED",

                SessionModel.session_date
                <= today,
            )
        )

        if student.join_date:

            session_query = (
                session_query.filter(
                    SessionModel.session_date
                    >= student.join_date
                )
            )

        sessions = (
            session_query
            .all()
        )

        conducted = len(
            sessions
        )

        session_ids = [
            session.id
            for session in sessions
        ]

        if session_ids:

            attended = (
                db.query(
                    func.count(
                        Attendance.id
                    )
                )
                .filter(
                    Attendance.student_id
                    == student.id,

                    Attendance.session_id.in_(
                        session_ids
                    ),

                    Attendance.status
                    == "Present",
                )
                .scalar()
                or 0
            )

        else:

            attended = 0

        absent = max(
            conducted - attended,
            0,
        )

        attendance_percentage = (
            round(
                (
                    attended
                    / conducted
                )
                * 100,
                2,
            )
            if conducted
            else 0
        )

        # -------------------------------------------------
        # ALL STUDENT PAYMENTS
        # -------------------------------------------------

        all_student_payments = (
            db.query(FeePayment)
            .filter(
                FeePayment.student_id
                == student.id
            )
            .order_by(
                FeePayment.payment_date.desc(),
                FeePayment.id.desc(),
            )
            .all()
        )

        # -------------------------------------------------
        # LIFECYCLE & TARGET MONTH FEE CALCULATION
        # -------------------------------------------------

        fee_data = calculate_student_lifecycle_fee(
            student=student,
            batch=batch,
            target_year=target_year,
            target_month=target_month,
            all_payments=all_student_payments,
            today=today,
        )

        curr_fee_status = fee_data["fee_status"]

        # Apply fee status filtering if requested
        if target_fee_filter:
            if target_fee_filter == "paid" and curr_fee_status != "PAID":
                continue
            elif target_fee_filter == "overdue" and curr_fee_status != "OVERDUE":
                continue
            elif target_fee_filter == "due_today" and curr_fee_status != "DUE TODAY":
                continue
            elif target_fee_filter == "unpaid" and curr_fee_status not in ("OVERDUE", "DUE TODAY", "UNPAID"):
                continue

        monthly_fee = fee_data["monthly_fee"]
        total_due_months = fee_data["total_due_months"]
        months_paid_count = fee_data["months_paid_count"]
        unpaid_months_count = fee_data["unpaid_months_count"]
        total_pending_amount = fee_data["total_pending_amount"]
        target_paid_amount = fee_data["target_paid_amount"]
        target_pending_amount = fee_data["target_pending_amount"]
        target_paid_date = fee_data["target_paid_date"]
        target_payment_method = fee_data["target_payment_method"]
        lifetime_total_paid = fee_data["lifetime_total_paid"]

        # -------------------------------------------------
        # PAYMENT HISTORY SHEET ROWS
        # -------------------------------------------------

        for payment_item in all_student_payments:

            payment_rows.append(
                {
                    "Payment ID":
                        payment_item.id,

                    "Student ID":
                        student.id,

                    "Student Name":
                        student.full_name,

                    "Batch Name":
                        batch.batch_name,

                    "Fee Month":
                        month_name(
                            payment_item.fee_month
                        ),

                    "Fee Year":
                        payment_item.fee_year,

                    "Amount":
                        int(
                            payment_item.net_payable
                            or 0
                        ),

                    "Payment Method":
                        str(
                            payment_item.payment_method
                        ),

                    "Payment Date":
                        format_date(
                            payment_item.payment_date
                        ),

                    "Status":
                        "PAID",
                }
            )

        # -------------------------------------------------
        # PDF STUDENT ITEM
        # -------------------------------------------------

        pdf_student_items.append(
            {
                "id":
                    student.id,
                "name":
                    student.full_name,
                "gender":
                    student.gender or "-",
                "age":
                    calculate_age(
                        student.dob
                    ),
                "dob":
                    format_date(
                        student.dob
                    ),
                "blood_group":
                    student.blood_group
                    or "",
                "batch_name":
                    batch.batch_name,
                "location":
                    batch.location
                    or "",
                "join_date":
                    format_date(
                        student.join_date
                    ),
                "parent_name":
                    student.parent_name or "-",
                "phone":
                    student.phone_number or "-",
                "status":
                    (
                        "ACTIVE"
                        if student.is_active
                        else "INACTIVE"
                    ),
                "attended":
                    attended,
                "conducted":
                    conducted,
                "monthly_fee":
                    monthly_fee,
                "fee_status":
                    curr_fee_status,
                "paid_amount":
                    target_paid_amount,
                "pending_amount":
                    target_pending_amount,
                "paid_date":
                    target_paid_date,
                "payment_method":
                    target_payment_method,
                "total_due_months":
                    total_due_months,
                "months_paid_count":
                    months_paid_count,
                "unpaid_months_count":
                    unpaid_months_count,
                "total_pending_amount":
                    total_pending_amount,
                "last_paid_amount":
                    fee_data.get("last_paid_amount", 0),
                "last_paid_month":
                    fee_data.get("last_paid_month", "-"),
            }
        )

        # -------------------------------------------------
        # STUDENT ROW (FOR EXCEL / CSV)
        # -------------------------------------------------

        rows.append(
            {
                "ID":
                    student.id,

                "Student Name":
                    student.full_name,

                "Gender":
                    student.gender or "",

                "Age":
                    calculate_age(
                        student.dob
                    ),

                "DOB":
                    format_date(
                        student.dob
                    ),

                "Blood":
                    student.blood_group
                    or "",

                "Join Date":
                    format_date(
                        student.join_date
                    ),

                "Parent Name":
                    student.parent_name or "",

                "Phone":
                    student.phone_number or "",

                "Status":
                    (
                        "ACTIVE"
                        if student.is_active
                        else "INACTIVE"
                    ),

                "Classes (Att/Cond)":
                    f"{attended}/{conducted}",

                "Monthly Fee":
                    monthly_fee,

                "Fee Status":
                    curr_fee_status,

                "Paid Amt":
                    target_paid_amount,

                "Pending":
                    target_pending_amount,

                "Paid Date":
                    target_paid_date,

                "Method":
                    target_payment_method,

                "Months (Paid/Due)":
                    f"{months_paid_count}/{total_due_months}",

                "Months Overdue":
                    unpaid_months_count,

                "Total Pending":
                    total_pending_amount,

                "Last Paid Amt":
                    fee_data.get("last_paid_amount", 0),

                "Last Paid Month":
                    fee_data.get("last_paid_month", "-"),
            }
        )

    if format == "pdf":

        if batch_id:

            batch_obj = (
                db.query(Batch)
                .filter(
                    Batch.id
                    == batch_id
                )
                .first()
            )

            selected_batch_name = (
                batch_obj.batch_name
                if batch_obj
                else f"Batch {batch_id}"
            )

            selected_location = (
                batch_obj.location
                if batch_obj
                and batch_obj.location
                else ""
            )

            is_multi = False
            batch_groups = None

        else:

            unique_batches = {}

            for item in pdf_student_items:

                b_name = item.get(
                    "batch_name",
                    "",
                )

                if b_name not in unique_batches:

                    unique_batches[b_name] = {
                        "location":
                            item.get(
                                "location",
                                "",
                            ),
                        "students": [],
                    }

                unique_batches[b_name][
                    "students"
                ].append(item)

            if len(unique_batches) <= 1:

                is_multi = False

                if unique_batches:

                    selected_batch_name = list(
                        unique_batches.keys()
                    )[0]

                    selected_location = (
                        unique_batches[
                            selected_batch_name
                        ]["location"]
                    )

                else:

                    selected_batch_name = (
                        "All Batches"
                    )

                    selected_location = ""

                batch_groups = None

            else:

                is_multi = True

                selected_batch_name = (
                    f"All Batches ({len(unique_batches)})"
                )

                locs = []
                for v in unique_batches.values():
                    loc_val = (v.get("location") or "").strip()
                    if loc_val and loc_val.lower() not in [l.lower() for l in locs]:
                        locs.append(loc_val)

                selected_location = (
                    ", ".join(locs)
                    if locs
                    else "All Locations"
                )

                batch_groups = unique_batches

        effective_filter_label = (
            f"{target_fee_filter.upper()} (Fee)"
            if target_fee_filter
            else (active_filter.upper() if active_filter else "ALL")
        )

        output = create_student_pdf(
            batch_name=selected_batch_name,
            location=selected_location,
            status_filter=effective_filter_label,
            students_data=pdf_student_items,
            is_multi_batch=is_multi,
            batches_data=batch_groups,
            month_label=target_month_label,
        )

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                    'attachment; filename="student-report.pdf"'
            },
        )

    return export_response(
        {
            "Students":
                rows,

            "Payment History":
                payment_rows,
        },
        format,
        "student-report",
        "Student Report",
    )


# =========================================================
# 4. BATCH EXPORT
# =========================================================


@router.get(
    "/batches/export"
)
def export_batch_report(

    batch_id: int | None = Query(
        default=None,
        gt=0,
        description="Filter by batch ID",
    ),

    month: int | None = Query(
        default=None,
        ge=1,
        le=12,
        description="Target fee month (1-12, defaults to current month)",
    ),

    year: int | None = Query(
        default=None,
        description="Target fee year (defaults to current year)",
    ),

    format: str = Query(
        default="xlsx",
        pattern="^(xlsx|pdf|csv)$",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):

    query = (
        db.query(Batch)
        .filter(
            Batch.is_active.is_(True)
        )
    )

    if batch_id:

        query = query.filter(
            Batch.id
            == batch_id
        )

    batches = (
        query
        .order_by(
            Batch.batch_name.asc()
        )
        .all()
    )

    rows = []
    pdf_batch_items = []

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year
    target_month_label = f"{month_name(target_month)} {target_year}"

    for batch in batches:

        # -------------------------------------------------
        # STUDENTS
        # -------------------------------------------------

        student_count = (
            db.query(
                func.count(
                    Student.id
                )
            )
            .filter(
                Student.batch_id
                == batch.id,

                Student.is_active.is_(True),
            )
            .scalar()
            or 0
        )

        # -------------------------------------------------
        # COMPLETED SESSIONS
        # -------------------------------------------------

        sessions = (
            db.query(SessionModel)
            .filter(
                SessionModel.batch_id
                == batch.id,

                SessionModel.status
                == "COMPLETED",

                SessionModel.session_date
                <= today,
            )
            .all()
        )

        conducted = len(
            sessions
        )

        session_ids = [
            session.id
            for session in sessions
        ]

        # -------------------------------------------------
        # PRESENT
        # -------------------------------------------------

        present = 0

        if session_ids:

            present = (
                db.query(
                    func.count(
                        Attendance.id
                    )
                )
                .filter(
                    Attendance.session_id.in_(
                        session_ids
                    ),

                    Attendance.status
                    == "Present",
                )
                .scalar()
                or 0
            )

        expected = (
            student_count
            * conducted
        )

        absent = max(
            expected - present,
            0,
        )

        attendance_percentage = (
            round(
                (
                    present
                    / expected
                )
                * 100,
                2,
            )
            if expected
            else 0
        )

        # -------------------------------------------------
        # TARGET MONTH REVENUE
        # -------------------------------------------------

        revenue = (
            db.query(
                func.coalesce(
                    func.sum(
                        FeePayment.net_payable
                    ),
                    0,
                )
            )
            .join(
                Student,
                FeePayment.student_id
                == Student.id,
            )
            .filter(
                Student.batch_id
                == batch.id,

                FeePayment.fee_month
                == target_month,

                FeePayment.fee_year
                == target_year,
            )
            .scalar()
            or 0
        )

        # -------------------------------------------------
        # PENDING FEES (TARGET MONTH)
        # -------------------------------------------------

        batch_students = (
            db.query(Student)
            .filter(
                Student.batch_id
                == batch.id,

                Student.is_active.is_(True),
            )
            .all()
        )

        pending_fee = 0

        for student in batch_students:

            paid = (
                db.query(FeePayment.id)
                .filter(
                    FeePayment.student_id
                    == student.id,

                    FeePayment.fee_month
                    == target_month,

                    FeePayment.fee_year
                    == target_year,
                )
                .first()
            )

            if not paid:

                pending_fee += int(
                    student.monthly_fee
                    or batch.monthly_fee
                    or 0
                )

        # -------------------------------------------------
        # PDF BATCH ITEM
        # -------------------------------------------------

        pdf_batch_items.append(
            {
                "id":
                    batch.id,
                "batch_name":
                    batch.batch_name,
                "location":
                    batch.location
                    or "",
                "level":
                    batch.level
                    or "",
                "class_type":
                    batch.class_type
                    or "",
                "training_days":
                    batch.training_days,
                "monthly_fee":
                    int(
                        batch.monthly_fee
                        or 0
                    ),
                "students":
                    student_count,
                "conducted":
                    conducted,
                "present":
                    present,
                "absent":
                    absent,
                "attendance_percentage":
                    attendance_percentage,
                "revenue":
                    int(revenue),
                "pending_fee":
                    pending_fee,
            }
        )

        # -------------------------------------------------
        # ROW (FOR EXCEL / CSV)
        # -------------------------------------------------

        rows.append(
            {
                "Batch ID":
                    batch.id,

                "Batch Name":
                    batch.batch_name,

                "Location":
                    batch.location
                    or "",

                "Level":
                    batch.level
                    or "",

                "Class Type":
                    batch.class_type
                    or "",

                "Training Days": (
                    ", ".join(batch.training_days)
                    if isinstance(
                        batch.training_days,
                        list,
                    )
                    else (
                        batch.training_days
                        or ""
                    )
                ),

                "Monthly Fee":
                    int(
                        batch.monthly_fee
                        or 0
                    ),

                "Students":
                    student_count,

                "Classes Conducted":
                    conducted,

                "Attendance Present":
                    present,

                "Attendance Absent":
                    absent,

                "Attendance %":
                    attendance_percentage,

                "Target Month":
                    target_month_label,

                "Month Revenue":
                    int(revenue),

                "Pending Fees":
                    pending_fee,
            }
        )

    if format == "pdf":

        batch_filter_name = None
        if batch_id and batches:
            batch_filter_name = batches[0].batch_name

        output = create_batch_pdf(
            batches_data=pdf_batch_items,
            batch_filter_name=batch_filter_name,
            month_label=target_month_label,
        )

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                    'attachment; filename="batch-report.pdf"'
            },
        )

    return export_response(
        {
            "Batches":
                rows
        },
        format,
        "batch-report",
        "Batch Report",
    )


# =========================================================
# 5. SINGLE STUDENT EXPORT
# =========================================================


@router.get(
    "/students/{student_id}/export"
)
def export_single_student_report(
    student_id: int,

    month: int | None = Query(
        default=None,
        ge=1,
        le=12,
        description="Target fee month (1-12, defaults to current month)",
    ),

    year: int | None = Query(
        default=None,
        description="Target fee year (defaults to current year)",
    ),

    format: str = Query(
        default="pdf",
        pattern="^(xlsx|pdf|csv)$",
        description="Export format: pdf | xlsx | csv",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):
    student = (
        db.query(Student)
        .filter(Student.id == student_id)
        .first()
    )

    if not student:
        raise HTTPException(
            status_code=404,
            detail=f"Student with ID {student_id} not found",
        )

    batch = (
        db.query(Batch)
        .filter(Batch.id == student.batch_id)
        .first()
    )

    today = date.today()
    target_month = month or today.month
    target_year = year or today.year
    target_month_label = f"{month_name(target_month)} {target_year}"

    # -------------------------------------------------
    # COMPLETED CLASSES & ATTENDANCE
    # -------------------------------------------------
    session_query = (
        db.query(SessionModel)
        .filter(
            SessionModel.batch_id == student.batch_id,
            SessionModel.status == "COMPLETED",
            SessionModel.session_date <= today,
        )
    )

    if student.join_date:
        session_query = session_query.filter(
            SessionModel.session_date >= student.join_date
        )

    sessions = (
        session_query
        .order_by(SessionModel.session_date.desc())
        .all()
    )

    conducted = len(sessions)
    session_ids = [session.id for session in sessions]

    attended = 0
    attended_session_ids = set()
    if session_ids:
        attendances = (
            db.query(Attendance)
            .filter(
                Attendance.student_id == student.id,
                Attendance.session_id.in_(session_ids),
                Attendance.status == "Present",
            )
            .all()
        )
        attended = len(attendances)
        attended_session_ids = {att.session_id for att in attendances}

    absent = max(conducted - attended, 0)
    attendance_percentage = (
        round((attended / conducted) * 100, 2)
        if conducted
        else 0
    )

    attendance_logs = []
    for s_item in sessions:
        is_pres = s_item.id in attended_session_ids
        attendance_logs.append(
            {
                "Session ID": s_item.id,
                "Session Date": format_date(s_item.session_date),
                "Topic / Note": getattr(s_item, "topic", None) or getattr(s_item, "notes", None) or "-",
                "Status": "PRESENT" if is_pres else "ABSENT",
            }
        )

    # -------------------------------------------------
    # PAYMENTS & LIFECYCLE LEDGER
    # -------------------------------------------------
    all_payments = (
        db.query(FeePayment)
        .filter(FeePayment.student_id == student.id)
        .order_by(
            FeePayment.payment_date.desc(),
            FeePayment.id.desc(),
        )
        .all()
    )

    lifecycle_data = calculate_student_lifecycle_fee(
        student=student,
        batch=batch,
        target_year=target_year,
        target_month=target_month,
        all_payments=all_payments,
        today=today,
    )

    # Group all payments by (fee_year, fee_month)
    paid_map: dict[tuple[int, int], dict] = {}
    for p in all_payments:
        amt = int(p.net_payable or 0)
        if p.fee_year and p.fee_month:
            k = (p.fee_year, p.fee_month)
            if k not in paid_map:
                paid_map[k] = {"total": 0, "payments": []}
            paid_map[k]["total"] += amt
            paid_map[k]["payments"].append(p)

    monthly_fee = lifecycle_data["monthly_fee"]
    join_date = student.join_date or (
        student.created_at.date()
        if getattr(student, "created_at", None)
        else date(target_year, target_month, 1)
    )

    monthly_ledger = []
    curr_y = join_date.year
    curr_m = join_date.month

    while (curr_y < target_year) or (curr_y == target_year and curr_m <= target_month):
        m_info = paid_map.get((curr_y, curr_m), {"total": 0, "payments": []})
        m_paid = m_info["total"]
        m_payments = m_info["payments"]
        m_due_date = date(curr_y, curr_m, 1)

        if (monthly_fee > 0 and m_paid >= monthly_fee) or (monthly_fee == 0 and (m_paid > 0 or len(m_payments) > 0)):
            m_status = "PAID"
            m_pending = 0
        else:
            m_pending = max(0, monthly_fee - m_paid)
            if today > m_due_date:
                m_status = "OVERDUE"
            elif today == m_due_date:
                m_status = "DUE TODAY"
            else:
                m_status = "UNPAID"

        latest_p = m_payments[0] if m_payments else None
        m_paid_date = format_date(latest_p.payment_date) if (latest_p and latest_p.payment_date) else "-"
        m_method = str(latest_p.payment_method) if (latest_p and latest_p.payment_method) else "-"

        monthly_ledger.append(
            {
                "Month & Year": f"{month_name(curr_m)} {curr_y}",
                "Fee Amount": monthly_fee,
                "Paid Amount": m_paid,
                "Pending Amount": m_pending,
                "Status": m_status,
                "Paid Date": m_paid_date,
                "Payment Method": m_method,
            }
        )

        if curr_m == 12:
            curr_y += 1
            curr_m = 1
        else:
            curr_m += 1

    # Payment transaction rows
    payment_history_rows = []
    for p_item in all_payments:
        payment_history_rows.append(
            {
                "Payment ID": p_item.id,
                "Payment Date": format_date(p_item.payment_date),
                "Fee Month": month_name(p_item.fee_month),
                "Fee Year": p_item.fee_year,
                "Amount": int(p_item.net_payable or 0),
                "Payment Method": str(p_item.payment_method),
                "Status": "PAID",
            }
        )

    clean_name = student.full_name.replace(" ", "_").lower()

    if format == "pdf":
        student_info = {
            "id": student.id,
            "name": student.full_name,
            "gender": student.gender or "-",
            "dob": format_date(student.dob),
            "age": calculate_age(student.dob),
            "blood_group": student.blood_group or "-",
            "join_date": format_date(student.join_date),
            "parent_name": student.parent_name or "-",
            "phone": student.phone_number or "-",
            "emergency_contact": student.emergency_contact or "-",
            "status": "ACTIVE" if student.is_active else "INACTIVE",
            "batch_name": batch.batch_name if batch else "-",
            "location": batch.location if (batch and batch.location) else "-",
            "monthly_fee": monthly_fee,
            "conducted": conducted,
            "attended": attended,
            "absent": absent,
            "attendance_percentage": attendance_percentage,
            "total_due_months": lifecycle_data["total_due_months"],
            "months_paid_count": lifecycle_data["months_paid_count"],
            "unpaid_months_count": lifecycle_data["unpaid_months_count"],
            "total_pending_amount": lifecycle_data["total_pending_amount"],
            "lifetime_total_paid": lifecycle_data["lifetime_total_paid"],
            "fee_status": lifecycle_data["fee_status"],
            "target_paid_amount": lifecycle_data["target_paid_amount"],
            "target_pending_amount": lifecycle_data["target_pending_amount"],
            "target_paid_date": lifecycle_data["target_paid_date"],
            "target_payment_method": lifecycle_data["target_payment_method"],
            "last_paid_amount": lifecycle_data["last_paid_amount"],
            "last_paid_month": lifecycle_data["last_paid_month"],
            "last_paid_date": lifecycle_data["last_paid_date"],
            "month_label": target_month_label,
        }

        output = create_single_student_pdf(
            student_data=student_info,
            monthly_ledger=monthly_ledger,
            payment_history=payment_history_rows,
        )

        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={
                "Content-Disposition":
                    f'attachment; filename="student-{student.id}-{clean_name}-report.pdf"'
            },
        )

    summary_profile_row = [
        {
            "Student ID": student.id,
            "Student Name": student.full_name,
            "Gender": student.gender or "-",
            "Age": calculate_age(student.dob),
            "DOB": format_date(student.dob),
            "Blood Group": student.blood_group or "-",
            "Join Date": format_date(student.join_date),
            "Parent Name": student.parent_name or "-",
            "Phone": student.phone_number or "-",
            "Emergency Contact": student.emergency_contact or "-",
            "Status": "ACTIVE" if student.is_active else "INACTIVE",
            "Batch Name": batch.batch_name if batch else "-",
            "Location": batch.location if (batch and batch.location) else "-",
            "Classes Conducted": conducted,
            "Classes Attended": attended,
            "Classes Absent": absent,
            "Attendance %": attendance_percentage,
            "Monthly Fee": monthly_fee,
            "Target Fee Month": target_month_label,
            "Fee Status": lifecycle_data["fee_status"],
            "Total Months (Since Join)": lifecycle_data["total_due_months"],
            "Months Paid": lifecycle_data["months_paid_count"],
            "Months Overdue": lifecycle_data["unpaid_months_count"],
            "Total Pending Balance": lifecycle_data["total_pending_amount"],
            "Lifetime Total Paid": lifecycle_data["lifetime_total_paid"],
            "Last Paid Amount": lifecycle_data["last_paid_amount"],
            "Last Paid Month": lifecycle_data["last_paid_month"],
            "Last Paid Date": lifecycle_data["last_paid_date"],
        }
    ]

    sheets = {
        "Profile & Summary": summary_profile_row,
        "Monthly Fee Ledger": monthly_ledger,
        "Payment Transactions": payment_history_rows,
        "Attendance Logs": attendance_logs,
    }

    return export_response(
        sheets,
        format,
        f"student-{student.id}-{clean_name}-report",
        f"Student Report - {student.full_name}",
    )


# =========================================================
# 6. FEE OVERVIEW EXPORT
# =========================================================


@router.get(
    "/fees/overview/export"
)
def export_fee_overview_report(
    month: int | None = Query(
        default=None,
        ge=1,
        le=12,
        description="Target fee month (1-12, defaults to current month)",
    ),

    year: int | None = Query(
        default=None,
        description="Target fee year (defaults to current year)",
    ),

    batch_id: int | None = Query(
        default=None,
        gt=0,
        description="Filter by batch ID",
    ),

    fee_status: str | None = Query(
        default=None,
        description="Filter by fee status: paid | overdue | due_today | unpaid | all",
    ),

    payment_method: str | None = Query(
        default=None,
        description="Filter by payment method: upi | cash | card | all",
    ),

    format: str = Query(
        default="xlsx",
        pattern="^(xlsx|pdf|csv)$",
        description="Export format: xlsx | pdf | csv",
    ),

    db: Session = Depends(get_db),

    current_admin: Admin = Depends(
        get_current_admin
    ),
):
    today = date.today()
    target_month = month or today.month
    target_year = year or today.year
    target_month_label = f"{month_name(target_month)} {target_year}"

    # Determine fee status filter
    effective_fee_status = (fee_status or "").strip().lower().replace(" ", "_")
    if effective_fee_status in ("all", ""):
        effective_fee_status = None

    # Determine payment method filter
    effective_pm_filter = (payment_method or "").strip().lower()
    if effective_pm_filter in ("all", ""):
        effective_pm_filter = None

    # 1. Batches Query
    batch_query = db.query(Batch).filter(Batch.is_active.is_(True))
    if batch_id:
        batch_query = batch_query.filter(Batch.id == batch_id)
    batches = batch_query.order_by(Batch.batch_name.asc()).all()
    batch_dict = {b.id: b for b in batches}

    # 2. Students Query
    student_query = (
        db.query(Student, Batch)
        .join(Batch, Student.batch_id == Batch.id)
        .filter(Student.is_active.is_(True), Batch.is_active.is_(True))
    )
    if batch_id:
        student_query = student_query.filter(Student.batch_id == batch_id)
    students = student_query.order_by(Batch.batch_name.asc(), Student.full_name.asc()).all()

    # 3. Monthly Fee Payments Query (for target month & year)
    payment_query = (
        db.query(FeePayment, Student, Batch)
        .join(Student, FeePayment.student_id == Student.id)
        .join(Batch, Student.batch_id == Batch.id)
        .filter(
            FeePayment.fee_month == target_month,
            FeePayment.fee_year == target_year,
            Student.is_active.is_(True),
            Batch.is_active.is_(True),
        )
    )
    if batch_id:
        payment_query = payment_query.filter(Student.batch_id == batch_id)
    if effective_pm_filter:
        payment_query = payment_query.filter(func.lower(FeePayment.payment_method) == effective_pm_filter)
    month_payments = payment_query.order_by(FeePayment.payment_date.desc()).all()

    # 4. Compute Student Dues & Lifecycle Stats
    student_rows = []
    paid_count = 0
    overdue_count = 0
    due_today_count = 0
    unpaid_count = 0
    total_expected_revenue = 0
    total_collected_revenue = 0
    total_pending_dues = 0

    for student, b_obj in students:
        # Fetch all payments for lifecycle fee calculation
        student_all_payments = (
            db.query(FeePayment)
            .filter(FeePayment.student_id == student.id)
            .order_by(FeePayment.payment_date.desc(), FeePayment.id.desc())
            .all()
        )

        lifecycle = calculate_student_lifecycle_fee(
            student=student,
            batch=b_obj,
            target_year=target_year,
            target_month=target_month,
            all_payments=student_all_payments,
            today=today,
        )

        st_status = lifecycle["fee_status"]
        monthly_fee = lifecycle["monthly_fee"]
        t_paid = lifecycle["target_paid_amount"]
        t_pending = lifecycle["target_pending_amount"]
        t_paid_date = lifecycle["target_paid_date"]
        t_method = lifecycle["target_payment_method"]

        total_expected_revenue += monthly_fee
        total_collected_revenue += t_paid
        total_pending_dues += t_pending

        if st_status == "PAID":
            paid_count += 1
        elif st_status == "OVERDUE":
            overdue_count += 1
        elif st_status == "DUE TODAY":
            due_today_count += 1
        else:
            unpaid_count += 1

        # Check filters
        if effective_fee_status:
            if effective_fee_status == "paid" and st_status != "PAID":
                continue
            elif effective_fee_status == "overdue" and st_status != "OVERDUE":
                continue
            elif effective_fee_status == "due_today" and st_status != "DUE TODAY":
                continue
            elif effective_fee_status == "unpaid" and st_status not in ("OVERDUE", "DUE TODAY", "UNPAID"):
                continue

        if effective_pm_filter:
            if t_method.lower() != effective_pm_filter:
                continue

        s_item = {
            "id": student.id,
            "name": student.full_name,
            "batch_name": b_obj.batch_name,
            "phone": student.phone_number or "-",
            "parent_name": student.parent_name or "-",
            "monthly_fee": monthly_fee,
            "fee_status": st_status,
            "paid_amount": t_paid,
            "pending_amount": t_pending,
            "paid_date": t_paid_date,
            "payment_method": t_method,
            "total_due_months": lifecycle["total_due_months"],
            "months_paid_count": lifecycle["months_paid_count"],
            "unpaid_months_count": lifecycle["unpaid_months_count"],
            "total_pending_amount": lifecycle["total_pending_amount"],
            "last_paid_amount": lifecycle["last_paid_amount"],
            "last_paid_month": lifecycle["last_paid_month"],
        }
        student_rows.append(s_item)

    total_students_count = len(students)
    collection_rate = (
        round((total_collected_revenue / total_expected_revenue) * 100, 2)
        if total_expected_revenue > 0
        else 0
    )

    # 5. Compute Batch-Level Performance
    batch_rows = []
    for b_id, b_obj in batch_dict.items():
        b_stu_list = [s for s, _ in students if s.batch_id == b_id]
        b_count = len(b_stu_list)
        b_m_fee = int(b_obj.monthly_fee or 0)
        b_expected = sum(
            int(s.monthly_fee if (s.monthly_fee is not None and s.monthly_fee > 0) else b_m_fee)
            for s in b_stu_list
        )

        b_paid_sum = (
            db.query(func.coalesce(func.sum(FeePayment.net_payable), 0))
            .join(Student, FeePayment.student_id == Student.id)
            .filter(
                Student.batch_id == b_id,
                FeePayment.fee_month == target_month,
                FeePayment.fee_year == target_year,
                Student.is_active.is_(True),
            )
            .scalar()
            or 0
        )
        b_collected = int(b_paid_sum)
        b_pending = max(0, b_expected - b_collected)
        b_recovery = round((b_collected / b_expected) * 100, 2) if b_expected > 0 else 0

        # Paid students in batch for target month
        b_paid_students_count = (
            db.query(func.count(func.distinct(FeePayment.student_id)))
            .join(Student, FeePayment.student_id == Student.id)
            .filter(
                Student.batch_id == b_id,
                FeePayment.fee_month == target_month,
                FeePayment.fee_year == target_year,
                Student.is_active.is_(True),
            )
            .scalar()
            or 0
        )

        batch_rows.append({
            "batch_name": b_obj.batch_name,
            "location": b_obj.location or "-",
            "students": b_count,
            "monthly_fee": b_m_fee,
            "expected": b_expected,
            "collected": b_collected,
            "pending": b_pending,
            "recovery_pct": b_recovery,
            "paid_vs_total": f"{b_paid_students_count} / {b_count}",
        })

    # 6. Compute Payment Methods Split
    pm_stats = {}
    for p_item, s_item, b_item in month_payments:
        pm_name = str(p_item.payment_method or "Other").strip().upper()
        amt = int(p_item.net_payable or 0)
        if pm_name not in pm_stats:
            pm_stats[pm_name] = {"count": 0, "amount": 0}
        pm_stats[pm_name]["count"] += 1
        pm_stats[pm_name]["amount"] += amt

    payment_methods_data = []
    for pm_name, pm_info in pm_stats.items():
        pm_share = (
            round((pm_info["amount"] / total_collected_revenue) * 100, 2)
            if total_collected_revenue > 0
            else 0
        )
        payment_methods_data.append({
            "method": pm_name,
            "count": pm_info["count"],
            "amount": pm_info["amount"],
            "share_pct": pm_share,
        })
    payment_methods_data.sort(key=lambda x: x["amount"], reverse=True)

    # 7. Payment Transactions Log
    transaction_rows = []
    for p_item, s_item, b_item in month_payments:
        transaction_rows.append({
            "Payment ID": p_item.id,
            "Payment Date": format_date(p_item.payment_date),
            "Student ID": s_item.id,
            "Student Name": s_item.full_name,
            "Batch Name": b_item.batch_name,
            "Fee Month": month_name(p_item.fee_month),
            "Fee Year": p_item.fee_year,
            "Amount": int(p_item.net_payable or 0),
            "Payment Method": str(p_item.payment_method),
            "Status": "PAID",
        })

    kpi_dict = {
        "total_expected_revenue": total_expected_revenue,
        "total_collected_revenue": total_collected_revenue,
        "total_pending_dues": total_pending_dues,
        "collection_rate": collection_rate,
        "total_students_count": total_students_count,
        "paid_students_count": paid_count,
        "overdue_students_count": overdue_count,
        "due_today_students_count": due_today_count,
        "unpaid_students_count": unpaid_count,
    }

    batch_filter_label = batch_dict[batch_id].batch_name if (batch_id and batch_id in batch_dict) else None

    # -------------------------------------------------
    # PDF EXPORT
    # -------------------------------------------------
    if format == "pdf":
        pdf_stream = create_fee_overview_pdf(
            month_label=target_month_label,
            kpi_data=kpi_dict,
            payment_methods_data=payment_methods_data,
            batches_data=batch_rows,
            students_data=student_rows,
            batch_filter_name=batch_filter_label,
            fee_status_filter=effective_fee_status,
            payment_method_filter=effective_pm_filter,
        )
        return StreamingResponse(
            pdf_stream,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="fee-overview-report-{target_month}-{target_year}.pdf"'
            },
        )

    # -------------------------------------------------
    # EXCEL / CSV DATASETS
    # -------------------------------------------------
    overview_summary_rows = [
        {
            "Metric": "Target Month",
            "Value": target_month_label,
        },
        {
            "Metric": "Total Expected Revenue",
            "Value": total_expected_revenue,
        },
        {
            "Metric": "Total Collected Revenue",
            "Value": total_collected_revenue,
        },
        {
            "Metric": "Total Pending Dues",
            "Value": total_pending_dues,
        },
        {
            "Metric": "Collection Recovery Rate (%)",
            "Value": f"{collection_rate:.1f}%",
        },
        {
            "Metric": "Total Active Students",
            "Value": total_students_count,
        },
        {
            "Metric": "Paid Students Count",
            "Value": paid_count,
        },
        {
            "Metric": "Overdue Students Count",
            "Value": overdue_count,
        },
        {
            "Metric": "Due Today / Unpaid Count",
            "Value": due_today_count + unpaid_count,
        },
    ]

    excel_batch_rows = [
        {
            "Batch Name": b["batch_name"],
            "Location": b["location"],
            "Students": b["students"],
            "Monthly Fee": b["monthly_fee"],
            "Expected Revenue": b["expected"],
            "Collected Revenue": b["collected"],
            "Pending Dues": b["pending"],
            "Recovery %": f"{b['recovery_pct']:.1f}%",
            "Paid / Total Students": b["paid_vs_total"],
        }
        for b in batch_rows
    ]

    excel_student_rows = [
        {
            "ID": s["id"],
            "Student Name": s["name"],
            "Batch Name": s["batch_name"],
            "Phone": s["phone"],
            "Parent Name": s["parent_name"],
            "Monthly Fee": s["monthly_fee"],
            "Target Month Status": s["fee_status"],
            "Paid Amount": s["paid_amount"],
            "Pending Amount": s["pending_amount"],
            "Paid Date": s["paid_date"],
            "Payment Method": s["payment_method"],
            "Months (Paid/Due)": f"{s['months_paid_count']}/{s['total_due_months']}",
            "Months Overdue": s["unpaid_months_count"],
            "Total Pending Balance": s["total_pending_amount"],
            "Last Paid Amount": s["last_paid_amount"],
            "Last Paid Month": s["last_paid_month"],
        }
        for s in student_rows
    ]

    excel_pm_rows = [
        {
            "Payment Method": pm["method"],
            "Transactions Count": pm["count"],
            "Amount Collected": pm["amount"],
            "% Share of Total": f"{pm['share_pct']:.1f}%",
        }
        for pm in payment_methods_data
    ]

    sheets = {
        "Executive Overview": overview_summary_rows,
        "Batch Performance": excel_batch_rows,
        "Student Fee Dues": excel_student_rows,
        "Payment Mode Summary": excel_pm_rows,
        "Transactions Log": transaction_rows,
    }

    return export_response(
        sheets,
        format,
        f"fee-overview-report-{target_month}-{target_year}",
        f"Fee Overview Report - {target_month_label}",
    )